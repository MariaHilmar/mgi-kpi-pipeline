import json
import unicodedata
from datetime import date, datetime
import re
import os
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet

try:
    import config as _config
except ImportError:
    _config = None

try:
    from detectar_area_funcional import AreaFuncionalDetector, build_detector
except ImportError:
    AreaFuncionalDetector = None
    build_detector = None

try:
    from excel_com_save import (
        DEV_GIT_EXCEL_COLOR,
        INFERRED_TIPO_EXCEL_COLOR,
        PROTECTED_SHEETS,
        save_workbook_preserving_filters,
    )
except ImportError:
    save_workbook_preserving_filters = None
    PROTECTED_SHEETS = frozenset({"Listas"})
    INFERRED_TIPO_EXCEL_COLOR = 49407
    DEV_GIT_EXCEL_COLOR = 15652797

try:
    from issue_filters import filtrar_issues_fechadas_antigas
except ImportError:
    filtrar_issues_fechadas_antigas = None

try:
    from inferir_tipo_issue import TipoIssueDetector, build_tipo_detector
except ImportError:
    TipoIssueDetector = None
    build_tipo_detector = None

try:
    from enriquecer_dev_git import DevGitInfo, GitDevEnricher, build_dev_enricher, resolve_desenvolvedor
except ImportError:
    DevGitInfo = None
    GitDevEnricher = None
    build_dev_enricher = None
    resolve_desenvolvedor = None

try:
    from atualizar_dashboard_kpis import atualizar_kpi_parceria
except ImportError:
    atualizar_kpi_parceria = None

try:
    from atualizar_graficos_dashboard import atualizar_graficos_novos
except ImportError:
    atualizar_graficos_novos = None

try:
    from atualizar_listas_taxonomia import sync_listas_taxonomia
except ImportError:
    sync_listas_taxonomia = None

try:
    from qualidade_dados import (
        atualizar_qualidade_dados,
        ensure_quality_columns,
        write_confidence,
    )
except ImportError:
    atualizar_qualidade_dados = None
    ensure_quality_columns = None
    write_confidence = None

try:
    from atualizar_releases_dashboard import atualizar_releases_dashboard
except ImportError:
    atualizar_releases_dashboard = None

try:
    from relatorio_excecoes import coletar_excecoes_wb, exportar as exportar_excecoes
except ImportError:
    coletar_excecoes_wb = None
    exportar_excecoes = None

try:
    from modulo_normalization import (
        apply_module_normalization,
        ensure_module_columns as ensure_module_cols,
    )
except ImportError:
    apply_module_normalization = None
    ensure_module_cols = None

try:
    from issue_keys import (
        DEFAULT_GITLAB_REPO,
        get_gitlab_repo,
        gitlab_work_item_url,
        lookup_issue,
        make_issue_key,
        make_key_from_parts,
        normalize_repo,
        parse_issue_key,
        repo_display_name,
        summarize_issues_by_repo,
    )
except ImportError:
    DEFAULT_GITLAB_REPO = "contratos_v2"

    def normalize_repo(raw):
        return raw or DEFAULT_GITLAB_REPO

    def repo_display_name(repo):
        return repo

    def gitlab_work_item_url(repo, iid):
        return f"https://gitlab.com/comprasnet/{repo}/-/work_items/{iid}"

    def lookup_issue(issues_by_id, issue_key):
        return issues_by_id.get(issue_key)

    def get_gitlab_repo(issue):
        return issue.get("gitlab_repo") or DEFAULT_GITLAB_REPO

    def make_issue_key(issue):
        return f"{get_gitlab_repo(issue)}:{str(issue.get('id', '')).strip()}"

    def make_key_from_parts(repo, iid):
        return f"{repo or DEFAULT_GITLAB_REPO}:{iid}"

# Destaque visual para tipo inferido (nao veio de label tipo:: no GitLab)
INFERRED_TIPO_FILL = PatternFill(fill_type="solid", fgColor="FFC000")
# Destaque azul claro para colunas enriquecidas via Git
DEV_GIT_FILL = PatternFill(fill_type="solid", fgColor="BDD7EE")

DEV_GIT_FIELDS: List[Tuple[str, str]] = [
    ("dev: tem branch", "Dev: Tem Branch"),
    ("dev: branch", "Dev: Branch"),
    ("dev: commits", "Dev: Commits"),
    ("dev: ultimo commit", "Dev: Último Commit"),
    ("dev: autor dev", "Dev: Autor Dev"),
    ("gitlab: mrs", "GitLab: MRs"),
    ("dev: mergeado?", "Dev: Mergeado?"),
]

DESENVOLVEDOR_COLUMN = ("desenvolvedor", "Desenvolvedor")

# Module standardization
MODULE_MAP = {
    'Gestão de Ata': 'Gestão de Atas',
    'GESTÃO DE ATAS': 'Gestão de Atas',
    'Transparência': 'Transparência',
    'Transparencia': 'Transparência',
    'API v2': 'API v2',
    'API': 'API v2',
    'Fiscalização': 'Fiscalização',
    'Fornecedor': 'Fornecedor',
    'Gestão Contratual': 'Gestão Contratual',
    'Gestão Financeira': 'Gestão Financeira',
    'Instrumento de Cobrança': 'Instrumento de Cobrança',
    'Jobs': 'Jobs',
    'Minuta de Empenho': 'Minuta de Empenho',
    'PNCP': 'PNCP',
    'Administração': 'Administração',
}

# Colunas preenchidas manualmente no Excel - nunca sobrescrever
MANUAL_ONLY_COLUMNS: Set[str] = {
    'Situação Análise',
    'Desenvolvedor Futuro',
    'Observação Geral',
    'Chamado',
    'Priorizar',
}

# Colunas sincronizadas do GitLab mas preservadas se ja preenchidas (issues existentes)
PROTECTED_COLUMNS: Set[str] = MANUAL_ONLY_COLUMNS | {
    'Módulo',
    'Área Funcional',
    'Estado',
    'Tipo',
    'Status',
    'Prioridade',
    'Equipe',
    'Parceria',
    'Sprint',
    'Assignee',
    'Autor',
    'Fechado em',
    'Lead Time',
    'Ano/Mês Criação',
    'Ano/Mês Fechamento',
    'Aberto?',
    'Fechado?',
    'Idade (dias)',
    'SLA > 90 dias',
    'Ano Criação',
    'Mês Criação Data',
    'Mês Fechamento Data',
}

_CALC_BAD_YEAR_CRITERION = re.compile(r'">=2024\)')
_CALC_GOOD_YEAR_CRITERION = '">="&2024)'


def _repair_calc_formulas(wb) -> int:
    """Corrige COUNTIFS com criterio de ano malformado na aba _Calc."""
    if "_Calc" not in wb.sheetnames:
        return 0
    ws = wb["_Calc"]
    fixed = 0
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if not isinstance(val, str) or not val.startswith("="):
                continue
            if _CALC_BAD_YEAR_CRITERION.search(val) and _CALC_GOOD_YEAR_CRITERION not in val:
                cell.value = _CALC_BAD_YEAR_CRITERION.sub(_CALC_GOOD_YEAR_CRITERION, val)
                fixed += 1
    if fixed:
        print(f"OK - {fixed} formulas corrigidas na aba _Calc (graficos Tipo/Prioridade/Modulo)")
    return fixed

# Nomes de coluna que o processador pode escrever (aliases normalizados)
WRITABLE_ALIASES: Dict[str, Set[str]] = {
    'id': {'#', 'id'},
    'title': {'titulo', 'título'},
    'module': {'modulo', 'módulo'},
    'area': {'area funcional', 'área funcional'},
    'created': {'criado em', 'data criacao', 'data criação'},
}

# Módulos permitidos para processamento (fallback se config indisponivel)
MODULOS_PERMITIDOS = {
    'Fiscalização',
    'Fornecedor'
}


def _modulo_permitido(module: str, all_modules: bool = False) -> bool:
    if all_modules:
        return True
    if _config is not None and hasattr(_config, 'modulo_permitido'):
        return _config.modulo_permitido(module)
    if not module:
        return False
    return module in MODULOS_PERMITIDOS


def _descricao_filtro_modulos(all_modules: bool = False) -> str:
    if all_modules:
        return "TODOS os modulos (MGI_ALL_MODULES=1)"
    if _config is not None and hasattr(_config, 'descricao_filtro_modulos'):
        return _config.descricao_filtro_modulos()
    return ", ".join(sorted(MODULOS_PERMITIDOS))

def parse_date(date_str: Optional[str]) -> Optional[datetime]:
    """Parse GitLab date format (formato humanizado ou ISO 8601)."""
    if not date_str:
        return None
    # Tenta primeiro o formato ISO (ex.: 2026-06-02T09:23:58)
    try:
        return datetime.fromisoformat(date_str.replace('Z', '+00:00').split('+')[0].strip())
    except (ValueError, AttributeError):
        pass
    # Formato humanizado (ex.: "Tuesday, June 2, 2026 at 9:23:58 AM GMT-3")
    try:
        clean = re.sub(r'(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday),\s*', '', date_str)
        clean = re.sub(r'\s+GMT[+-]\d+', '', clean)
        return datetime.strptime(clean, '%B %d, %Y at %I:%M:%S %p')
    except (ValueError, TypeError):
        return None

def extract_module(title: str) -> str:
    """Extrai modulo do titulo e normaliza para um dos 12 canonicos quando possivel."""
    from taxonomy import extract_module_tag, normalize_module_to_canonical

    tag = extract_module_tag(title)
    if not tag:
        return ""
    canon = normalize_module_to_canonical(tag)
    if canon:
        return canon
    return MODULE_MAP.get(tag, tag)

def extract_functional_area(title: str) -> str:
    """Extract functional area from title (only when declared in parentheses)."""
    match = re.search(r'\]\s*\(([^)]+)\)', title)
    if not match:
        return ''
    area = match.group(1).strip()
    if area.startswith('- '):
        area = area[2:]
    return area


def _resolve_area_detection(
    issue: Dict,
    title: str,
    area_detector: Optional["AreaFuncionalDetector"],
):
    """Retorna AreaDetection com area e confianca."""
    title_area = extract_functional_area(title)
    if area_detector is None:
        if title_area:
            from detectar_area_funcional import AreaDetection

            return AreaDetection(title_area, "titulo_explicito", 1.0)
        from detectar_area_funcional import AreaDetection

        return AreaDetection("", "none", 0.0)
    return area_detector.detect(issue, title_area=title_area)


def _resolve_functional_area(
    issue: Dict,
    title: str,
    area_detector: Optional["AreaFuncionalDetector"],
) -> str:
    """Resolve area funcional: titulo explicito, codigo Git ou vazio."""
    return _resolve_area_detection(issue, title, area_detector).area


def _backfill_areas_for_sheet(
    ws: Worksheet,
    columns: Dict[str, int],
    data_start_row: int,
    id_col: int,
    repo_col: Optional[int],
    issues_by_id: Dict[str, Dict],
    area_detector: Optional["AreaFuncionalDetector"],
    force: bool = False,
    only_area: Optional[str] = None,
) -> int:
    """Preenche ou corrige Area Funcional em todas as linhas com issue conhecida."""
    if not area_detector or "area" not in columns:
        return 0

    area_col = columns["area"]
    filled = 0
    total = ws.max_row - data_start_row + 1
    for row in range(data_start_row, ws.max_row + 1):
        if (row - data_start_row) % 500 == 0 and row > data_start_row:
            print(f"   ... {row - data_start_row}/{total} linhas", flush=True)

        issue_key = _row_issue_key(ws, row, id_col, repo_col)
        if not issue_key:
            continue
        issue = lookup_issue(issues_by_id, issue_key)
        if not issue:
            continue

        current = ws.cell(row=row, column=area_col).value
        current_str = str(current).strip() if current not in (None, "") else ""

        if only_area:
            if not force or current_str.casefold() != only_area.strip().casefold():
                continue
        elif not force and current not in (None, ""):
            continue

        title = issue.get("title", "")
        area = _resolve_functional_area(issue, title, area_detector)
        if area and (force or only_area or current in (None, "")):
            ws.cell(row=row, column=area_col).value = area
            filled += 1

    return filled


def _map_estado(state: Optional[str]) -> str:
    if state == 'opened':
        return 'Aberto'
    if state == 'closed':
        return 'Fechado'
    return state or ''


def _parse_labels(labels: List[str]) -> Dict[str, str]:
    parsed = {
        'tipo': '', 'status': '', 'equipe': '', 'parceria': '',
        'prioridade': '', 'solicitante': '', 'alteracao_escopo': 'Não',
    }
    for label in labels or []:
        if label.startswith('tipo::'):
            parsed['tipo'] = label.split('::', 1)[1]
        elif label.startswith('status::'):
            parsed['status'] = label.split('::', 1)[1]
        elif label.startswith('Equipe::'):
            parsed['equipe'] = label.split('::', 1)[1]
        elif label.startswith('Parceria::'):
            parsed['parceria'] = label.split('::', 1)[1]
        elif label.startswith('priority::'):
            parsed['prioridade'] = label.split('::', 1)[1]
        elif label.startswith('Solicitante::'):
            parsed['solicitante'] = label.split('::', 1)[1]
        elif label.strip() == 'Alteração Escopo':
            parsed['alteracao_escopo'] = 'Sim'
    return parsed


def _format_assignees(issue: Dict) -> str:
    assignees = issue.get('assignees') or []
    names = [person.get('name', '') for person in assignees if person.get('name')]
    return ', '.join(names)


def _build_full_header_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        norm = _normalize_header(ws.cell(row=header_row, column=col).value)
        if norm and norm not in header_map:
            header_map[norm] = col
    return header_map


def _set_cell(
    ws: Worksheet,
    header_map: Dict[str, int],
    field: str,
    row: int,
    value,
    force: bool = False,
) -> None:
    col = header_map.get(field)
    if col is None or value in (None, ''):
        return
    current = ws.cell(row=row, column=col).value
    if not force and current not in (None, ''):
        return
    ws.cell(row=row, column=col).value = value


def _row_needs_gitlab_metadata(ws: Worksheet, row: int, header_map: Dict[str, int]) -> bool:
    estado_col = header_map.get('estado')
    if not estado_col:
        return True
    value = ws.cell(row=row, column=estado_col).value
    return value is None or str(value).strip() == ''


def _sync_label_columns(
    ws: Worksheet,
    row: int,
    header_map: Dict[str, int],
    issue: Dict,
) -> None:
    """Sincroniza colunas derivadas de labels GitLab (sempre, inclusive linhas existentes)."""
    labels = _parse_labels(issue.get('labels') or [])
    _set_cell(ws, header_map, 'prioridade', row, labels['prioridade'], force=True)
    _set_cell(ws, header_map, 'solicitante', row, labels['solicitante'], force=True)
    _set_cell(ws, header_map, 'alteracao escopo', row, labels['alteracao_escopo'], force=True)


def _sync_gitlab_metadata(
    ws: Worksheet,
    row: int,
    header_map: Dict[str, int],
    issue: Dict,
    created_date: Optional[datetime],
) -> None:
    """Preenche colunas derivadas do GitLab (labels, estado, datas, metricas)."""
    labels = _parse_labels(issue.get('labels') or [])
    estado = _map_estado(issue.get('state', ''))
    closed_date = parse_date(issue.get('closedDate', ''))
    milestone = (issue.get('milestone') or {}).get('title', '')

    _sync_label_columns(ws, row, header_map, issue)
    _set_cell(ws, header_map, 'estado', row, estado, force=True)
    _set_cell(ws, header_map, 'status', row, labels['status'], force=True)
    _set_cell(ws, header_map, 'equipe', row, labels['equipe'], force=True)
    _set_cell(ws, header_map, 'parceria', row, labels['parceria'], force=True)
    _set_cell(ws, header_map, 'sprint', row, milestone, force=True)
    _set_cell(ws, header_map, 'assignee', row, _format_assignees(issue), force=True)
    _set_cell(ws, header_map, 'autor', row, (issue.get('author') or {}).get('name', ''), force=True)

    if created_date:
        ym_created = f"{created_date.year}/{created_date.month:02d}"
        _set_cell(ws, header_map, 'ano/mes criacao', row, ym_created, force=True)
        _set_cell(ws, header_map, 'ano criacao', row, created_date.year, force=True)
        _set_cell(ws, header_map, 'mes criacao data', row, date(created_date.year, created_date.month, 1), force=True)

    if closed_date:
        _set_cell(ws, header_map, 'fechado em', row, closed_date, force=True)
        ym_closed = f"{closed_date.year}/{closed_date.month:02d}"
        _set_cell(ws, header_map, 'ano/mes fechamento', row, ym_closed, force=True)
        _set_cell(ws, header_map, 'mes fechamento data', row, date(closed_date.year, closed_date.month, 1), force=True)
        if created_date:
            lead_days = max((closed_date.date() - created_date.date()).days, 0)
            _set_cell(ws, header_map, 'lead time', row, lead_days, force=True)

    aberto = estado == 'Aberto'
    fechado = estado == 'Fechado'
    _set_cell(ws, header_map, 'aberto?', row, 'Sim' if aberto else 'Não', force=True)
    _set_cell(ws, header_map, 'fechado?', row, 'Sim' if fechado else 'Não', force=True)

    if aberto and created_date:
        idade = max((date.today() - created_date.date()).days, 0)
        _set_cell(ws, header_map, 'idade (dias)', row, idade, force=True)
        _set_cell(ws, header_map, 'sla > 90 dias', row, 'Sim' if idade > 90 else 'Não', force=True)
    elif fechado:
        _set_cell(ws, header_map, 'idade (dias)', row, 0, force=True)
        _set_cell(ws, header_map, 'sla > 90 dias', row, 'Não', force=True)


def _sync_tipo_column(
    ws: Worksheet,
    row: int,
    header_map: Dict[str, int],
    issue: Dict,
    tipo_detector: Optional["TipoIssueDetector"],
    inferred_tipo_ids: Set[str],
) -> bool:
    """Preenche Tipo a partir da label ou inferencia Git/titulo. Retorna True se inferido."""
    tipo_col = header_map.get('tipo')
    if not tipo_col:
        return False

    labels = _parse_labels(issue.get('labels') or [])
    issue_key = make_issue_key(issue)
    cell = ws.cell(row=row, column=tipo_col)

    if labels['tipo']:
        cell.value = labels['tipo']
        if issue_key:
            inferred_tipo_ids.discard(issue_key)
        return False

    inferred = ''
    if tipo_detector:
        inferred = tipo_detector.detect(issue).tipo

    if inferred:
        cell.value = inferred
        if issue_key:
            inferred_tipo_ids.add(issue_key)
        return True

    cell.value = ''
    if issue_key:
        inferred_tipo_ids.discard(issue_key)
    return False


def _row_issue_key(
    ws: Worksheet,
    row: int,
    id_col: int,
    repo_col: Optional[int],
) -> Optional[str]:
    """Monta chave composta repositorio:iid a partir da linha da planilha."""
    iid = _normalize_id(ws.cell(row=row, column=id_col).value)
    if not iid:
        return None
    repo = DEFAULT_GITLAB_REPO
    if repo_col:
        raw = ws.cell(row=row, column=repo_col).value
        if raw is not None and str(raw).strip():
            repo = normalize_repo(str(raw).strip())
    return make_key_from_parts(repo, iid)


def _apply_inferred_tipo_highlights(
    ws: Worksheet,
    header_map: Dict[str, int],
    data_start_row: int,
    id_col: int,
    repo_col: Optional[int],
    inferred_tipo_ids: Set[str],
    processed_issue_ids: Set[str],
) -> List[Tuple[int, int, int]]:
    """Aplica fundo laranja na coluna Tipo para valores inferidos."""
    tipo_col = header_map.get('tipo')
    if not tipo_col:
        return []

    highlights: List[Tuple[int, int, int]] = []
    clear_fill = PatternFill(fill_type=None)

    for row in range(data_start_row, ws.max_row + 1):
        issue_key = _row_issue_key(ws, row, id_col, repo_col)
        if not issue_key:
            continue

        cell = ws.cell(row=row, column=tipo_col)
        if issue_key in inferred_tipo_ids and cell.value:
            cell.fill = INFERRED_TIPO_FILL
            highlights.append((row, tipo_col, INFERRED_TIPO_EXCEL_COLOR))
        elif issue_key in processed_issue_ids:
            cell.fill = clear_fill

    return highlights


def _fill_missing_tipos_in_sheet(
    ws: Worksheet,
    header_map: Dict[str, int],
    data_start_row: int,
    id_col: int,
    repo_col: Optional[int],
    issues_by_id: Dict[str, Dict],
    tipo_detector: Optional["TipoIssueDetector"],
    inferred_tipo_ids: Set[str],
) -> int:
    """Completa Tipo vazio em linhas ja existentes na planilha."""
    tipo_col = header_map.get('tipo')
    if not tipo_col or not tipo_detector:
        return 0

    filled = 0
    for row in range(data_start_row, ws.max_row + 1):
        issue_key = _row_issue_key(ws, row, id_col, repo_col)
        if not issue_key:
            continue

        current_tipo = ws.cell(row=row, column=tipo_col).value
        if current_tipo not in (None, ''):
            continue

        issue = lookup_issue(issues_by_id, issue_key)
        if not issue:
            continue

        if _sync_tipo_column(ws, row, header_map, issue, tipo_detector, inferred_tipo_ids):
            filled += 1

    return filled


EXTRA_LABEL_FIELDS: List[tuple] = [
    ('solicitante', 'Solicitante'),
    ('alteracao escopo', 'Alteração Escopo'),
]

REPOSITORIO_FIELD = ('repositorio', 'Repositório')


def _ensure_repositorio_column(
    ws: Worksheet,
    header_row: int,
    header_map: Dict[str, int],
) -> Dict[str, int]:
    """Garante coluna Repositório (contratos_v2 | contratos) na aba Dados."""
    field_key, title = REPOSITORIO_FIELD
    if field_key in header_map:
        return header_map
    next_col = ws.max_column + 1
    ws.cell(row=header_row, column=next_col).value = title
    header_map[field_key] = next_col
    return header_map


def _ensure_extra_label_columns(
    ws: Worksheet,
    header_row: int,
    header_map: Dict[str, int],
) -> Dict[str, int]:
    """Garante colunas de Solicitante e Alteração Escopo no cabeçalho da aba Dados."""
    next_col = ws.max_column + 1
    for field_key, title in EXTRA_LABEL_FIELDS:
        if field_key in header_map:
            continue
        ws.cell(row=header_row, column=next_col).value = title
        header_map[field_key] = next_col
        next_col += 1
    return header_map


def _ensure_dev_git_columns(
    ws: Worksheet,
    header_row: int,
    header_map: Dict[str, int],
) -> Dict[str, int]:
    """Garante colunas de enriquecimento Git no cabecalho da aba Dados."""
    next_col = ws.max_column + 1
    for field_key, title in DEV_GIT_FIELDS:
        if field_key in header_map:
            continue
        ws.cell(row=header_row, column=next_col).value = title
        header_map[field_key] = next_col
        next_col += 1
    field_key, title = DESENVOLVEDOR_COLUMN
    if field_key not in header_map:
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if header and str(header).strip().lower() == title.lower():
                header_map[field_key] = col
                break
    if field_key not in header_map:
        ws.cell(row=header_row, column=next_col).value = title
        header_map[field_key] = next_col
    return header_map


def _write_dev_git_row(
    ws: Worksheet,
    row: int,
    header_map: Dict[str, int],
    info: "DevGitInfo",
    issue: Optional[Dict] = None,
) -> None:
    values = {
        "dev: tem branch": info.tem_branch,
        "dev: branch": info.branch,
        "dev: commits": info.commits,
        "dev: ultimo commit": info.ultimo_commit,
        "dev: autor dev": info.autor_dev,
        "gitlab: mrs": info.mr_gitlab,
        "dev: mergeado?": info.mergeado,
    }
    for field_key, value in values.items():
        col = header_map.get(field_key)
        if col is None:
            continue
        ws.cell(row=row, column=col).value = value

    dev_col = header_map.get("desenvolvedor")
    if dev_col and issue is not None and resolve_desenvolvedor:
        ws.cell(row=row, column=dev_col).value = resolve_desenvolvedor(issue, info)


def _enrich_dev_git_for_sheet(
    ws: Worksheet,
    header_map: Dict[str, int],
    data_start_row: int,
    id_col: int,
    repo_col: Optional[int],
    issues_by_id: Dict[str, Dict],
    dev_enricher: Optional["GitDevEnricher"],
) -> Tuple[List[Tuple[int, int, int]], int, int]:
    """Preenche colunas Dev/Git para todas as linhas com issue conhecida."""
    if not dev_enricher:
        return [], 0, 0

    highlights: List[Tuple[int, int, int]] = []
    enriched = 0
    with_branch = 0

    for row in range(data_start_row, ws.max_row + 1):
        issue_key = _row_issue_key(ws, row, id_col, repo_col)
        if not issue_key:
            continue
        issue = lookup_issue(issues_by_id, issue_key)
        if not issue:
            continue

        info = dev_enricher.enrich(issue)
        _write_dev_git_row(ws, row, header_map, info, issue)
        enriched += 1
        if info.tem_branch == "Sim":
            with_branch += 1

        dev_col = header_map.get("desenvolvedor")
        highlight_dev = info.tem_branch == "Sim" or info.commits > 0 or info.mergeado == "Sim"
        if highlight_dev:
            for field_key, _ in DEV_GIT_FIELDS:
                col = header_map.get(field_key)
                if not col:
                    continue
                ws.cell(row=row, column=col).fill = DEV_GIT_FILL
                highlights.append((row, col, DEV_GIT_EXCEL_COLOR))
            if dev_col and resolve_desenvolvedor and resolve_desenvolvedor(issue, info):
                ws.cell(row=row, column=dev_col).fill = DEV_GIT_FILL
                highlights.append((row, dev_col, DEV_GIT_EXCEL_COLOR))

    return highlights, enriched, with_branch


def _backfill_desenvolvedor_for_sheet(
    ws: Worksheet,
    header_map: Dict[str, int],
    data_start_row: int,
    id_col: int,
    repo_col: Optional[int],
    issues_by_id: Dict[str, Dict],
    dev_enricher: Optional["GitDevEnricher"],
    force: bool = False,
) -> int:
    """Preenche coluna Desenvolvedor a partir do Git e assignee GitLab."""
    if not dev_enricher or not resolve_desenvolvedor:
        return 0

    dev_col = header_map.get("desenvolvedor")
    if not dev_col:
        return 0

    filled = 0
    total = ws.max_row - data_start_row + 1
    for row in range(data_start_row, ws.max_row + 1):
        if (row - data_start_row) % 500 == 0 and row > data_start_row:
            print(f"   ... {row - data_start_row}/{total} linhas", flush=True)

        issue_key = _row_issue_key(ws, row, id_col, repo_col)
        if not issue_key:
            continue
        issue = lookup_issue(issues_by_id, issue_key)
        if not issue:
            continue

        current = ws.cell(row=row, column=dev_col).value
        if not force and current not in (None, ""):
            continue

        info = dev_enricher.enrich(issue)
        dev = resolve_desenvolvedor(issue, info)
        if dev:
            ws.cell(row=row, column=dev_col).value = dev
            autor_col = header_map.get("dev: autor dev")
            if autor_col and info.autor_dev:
                ws.cell(row=row, column=autor_col).value = info.autor_dev
            filled += 1

    return filled


def _normalize_header(name: Optional[str]) -> str:
    """Normaliza nome de coluna para comparacao (sem acentos, minusculo)."""
    if not name:
        return ''
    text = str(name).strip().lower()
    return ''.join(
        c for c in unicodedata.normalize('NFD', text)
        if unicodedata.category(c) != 'Mn'
    )


def _normalize_id(value) -> Optional[str]:
    """Normaliza ID da issue para comparacao str/int."""
    if value is None:
        return None
    if isinstance(value, float):
        return str(int(value))
    text = str(value).strip()
    if not text or text == '#':
        return None
    if text.isdigit():
        return text
    return text


def _resolve_sheet_layout(ws: Worksheet) -> Tuple[int, int, Dict[str, int]]:
    """Detecta linha de cabecalho e mapeia colunas da aba Dados.

    O MGI_Dashboard usa cabecalhos na linha 2 e dados a partir da linha 3.
    Retorna: (header_row, data_start_row, columns) onde columns mapeia
    chaves logicas ('id', 'title', ...) para numero da coluna.
    """
    header_row = 1
    for row in range(1, min(11, ws.max_row + 1)):
        labels = {
            _normalize_header(ws.cell(row=row, column=col).value)
            for col in range(1, ws.max_column + 1)
        }
        labels.discard('')
        if 'titulo' in labels or 'modulo' in labels:
            header_row = row
            break

    columns: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=col).value
        norm = _normalize_header(raw)
        if not norm:
            continue
        for field, aliases in WRITABLE_ALIASES.items():
            if norm in aliases and field not in columns:
                columns[field] = col

    data_start_row = header_row + 1
    return header_row, data_start_row, columns


ISSUES_TABLE_NAME = 'tbIssues'


def _expand_issues_table(ws: Worksheet) -> Optional[str]:
    """Expande a tabela estruturada tbIssues ate a ultima linha com dados."""
    if ISSUES_TABLE_NAME not in ws.tables:
        return None

    table = ws.tables[ISSUES_TABLE_NAME]
    min_col, min_row, max_col, old_max_row = range_boundaries(table.ref)
    new_max_row = ws.max_row
    if new_max_row <= old_max_row:
        return table.ref

    new_ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{new_max_row}"
    )
    table.ref = new_ref
    if table.autoFilter is not None:
        table.autoFilter.ref = new_ref
    return new_ref


def _fill_derived_formulas(ws: Worksheet, row: int, header_map: Dict[str, int]) -> None:
    """Preenche formulas apenas quando a linha ainda nao tem metricas estaticas."""
    idade_col = header_map.get('idade (dias)')
    if idade_col and ws.cell(row=row, column=idade_col).value is not None:
        return

    created_col = header_map.get('criado em')
    estado_col = header_map.get('estado')
    if not created_col or not estado_col:
        return

    created_letter = get_column_letter(created_col)
    estado_letter = get_column_letter(estado_col)
    idade_letter = get_column_letter(idade_col) if idade_col else 'U'

    ym_col = header_map.get('ano/mes criacao')
    if ym_col and not ws.cell(row=row, column=ym_col).value:
        ws.cell(row=row, column=ym_col).value = f'=TEXT({created_letter}{row},"YYYY/MM")'

    aberto_col = header_map.get('aberto?')
    if aberto_col and not ws.cell(row=row, column=aberto_col).value:
        ws.cell(row=row, column=aberto_col).value = f'=IF({estado_letter}{row}="Aberto","Sim","Não")'

    fechado_col = header_map.get('fechado?')
    if fechado_col and not ws.cell(row=row, column=fechado_col).value:
        ws.cell(row=row, column=fechado_col).value = f'=IF({estado_letter}{row}="Fechado","Sim","Não")'

    if idade_col and not ws.cell(row=row, column=idade_col).value:
        ws.cell(row=row, column=idade_col).value = f'=IF({estado_letter}{row}="Aberto",TODAY()-{created_letter}{row},0)'

    sla_col = header_map.get('sla > 90 dias')
    if sla_col and not ws.cell(row=row, column=sla_col).value:
        ws.cell(row=row, column=sla_col).value = (
            f'=IF(AND({estado_letter}{row}="Aberto",{idade_letter}{row}>90),"Sim","Não")'
        )


def _issue_id_sort_key(value) -> int:
    normalized = _normalize_id(value)
    if normalized and normalized.isdigit():
        return int(normalized)
    return -1


def _sort_issue_rows_desc(
    ws: Worksheet,
    data_start_row: int,
    id_col: int,
    repo_col: Optional[int] = None,
) -> int:
    """Reordena linhas por # decrescente (desempate: repositorio)."""
    max_col = ws.max_column
    rows_data: List[List] = []

    for row in range(data_start_row, ws.max_row + 1):
        issue_id = ws.cell(row=row, column=id_col).value
        if issue_id is None or str(issue_id).strip() in ('', '#'):
            continue
        rows_data.append([
            ws.cell(row=row, column=col).value
            for col in range(1, max_col + 1)
        ])

    if len(rows_data) < 2:
        return len(rows_data)

    def sort_key(values: List) -> Tuple:
        repo_val = values[repo_col - 1] if repo_col and len(values) >= repo_col else ""
        return (
            _issue_id_sort_key(values[id_col - 1]),
            normalize_repo(str(repo_val or DEFAULT_GITLAB_REPO)),
        )

    rows_data.sort(key=sort_key, reverse=True)

    for idx, row_values in enumerate(rows_data):
        target_row = data_start_row + idx
        for col, value in enumerate(row_values, start=1):
            ws.cell(row=target_row, column=col).value = value

    last_data_row = data_start_row + len(rows_data) - 1
    for row in range(last_data_row + 1, ws.max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).value = None

    return len(rows_data)


def _remove_rows_not_in_active_set(
    ws: Worksheet,
    data_start_row: int,
    id_col: int,
    repo_col: Optional[int],
    active_ids: Set[str],
) -> int:
    """Remove linhas cuja issue nao faz parte do conjunto ativo (carga inicial)."""
    rows_to_delete: List[int] = []
    for row in range(data_start_row, ws.max_row + 1):
        issue_key = _row_issue_key(ws, row, id_col, repo_col)
        if issue_key and issue_key not in active_ids:
            rows_to_delete.append(row)

    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row, 1)

    return len(rows_to_delete)


def _finalize_sheet_after_inserts(
    ws: Worksheet,
    data_start_row: int,
    header_map: Dict[str, int],
) -> Tuple[Optional[str], int]:
    """Expande tbIssues e aplica formulas residuais nas linhas incompletas."""
    formulas_applied = 0
    for row in range(data_start_row, ws.max_row + 1):
        issue_id = ws.cell(row=row, column=1).value
        if issue_id is None or str(issue_id).strip() in ('', '#'):
            continue
        if _row_needs_gitlab_metadata(ws, row, header_map):
            continue
        idade_col = header_map.get('idade (dias)')
        if idade_col and ws.cell(row=row, column=idade_col).value is None:
            _fill_derived_formulas(ws, row, header_map)
            formulas_applied += 1

    new_ref = _expand_issues_table(ws)
    return new_ref, formulas_applied


def _set_issue_id_cell(
    ws: Worksheet,
    row: int,
    id_col: int,
    issue_id: str,
    repo: str,
) -> None:
    """Grava # com hyperlink para o work item no GitLab."""
    from openpyxl.styles import Font

    cell = ws.cell(row=row, column=id_col)
    cell.value = int(issue_id) if str(issue_id).isdigit() else issue_id
    cell.hyperlink = gitlab_work_item_url(repo, str(issue_id))
    cell.font = Font(color="0563C1", underline="single")


def _row_filled_score(ws: Worksheet, row: int, max_col: int) -> int:
    score = 0
    for col in range(1, max_col + 1):
        if ws.cell(row=row, column=col).value not in (None, ""):
            score += 1
    return score


def _issues_for_iid(issues_by_id: Dict[str, Dict], iid: str) -> List[Tuple[str, Dict]]:
    suffix = f":{iid}"
    return [(k, issues_by_id[k]) for k in issues_by_id if k.endswith(suffix)]


def _match_issue_for_row(
    ws: Worksheet,
    row: int,
    id_col: int,
    repo_col: Optional[int],
    title_col: Optional[int],
    issues_by_id: Dict[str, Dict],
) -> Tuple[Optional[Dict], Optional[str]]:
    """Resolve issue JSON para a linha (chave exata ou unica candidata)."""
    issue_key = _row_issue_key(ws, row, id_col, repo_col)
    if not issue_key:
        return None, None

    issue = issues_by_id.get(issue_key)
    if issue:
        return issue, issue_key

    iid = parse_issue_key(issue_key)[1]
    candidates = _issues_for_iid(issues_by_id, iid)
    if len(candidates) == 1:
        key, issue = candidates[0]
        return issue, key

    if len(candidates) > 1 and title_col:
        title = ws.cell(row=row, column=title_col).value
        if title:
            for key, cand in candidates:
                if cand.get("title") == title:
                    return cand, key
    return None, issue_key


def _dedupe_dados_rows(
    ws: Worksheet,
    data_start_row: int,
    id_col: int,
    repo_col: Optional[int],
    title_col: Optional[int],
    issues_by_id: Dict[str, Dict],
) -> int:
    """Remove linhas duplicadas (mesmo repositorio:#) e linhas com repo errado."""
    max_col = ws.max_column
    rows_to_delete: Set[int] = set()

    key_rows: Dict[str, List[int]] = {}
    for row in range(data_start_row, ws.max_row + 1):
        key = _row_issue_key(ws, row, id_col, repo_col)
        if key:
            key_rows.setdefault(key, []).append(row)

    for rows in key_rows.values():
        if len(rows) <= 1:
            continue
        rows.sort(key=lambda r: _row_filled_score(ws, r, max_col), reverse=True)
        rows_to_delete.update(rows[1:])

    resolved_key_rows: Dict[str, int] = {}
    for row in range(data_start_row, ws.max_row + 1):
        if row in rows_to_delete:
            continue
        issue, resolved_key = _match_issue_for_row(
            ws, row, id_col, repo_col, title_col, issues_by_id
        )
        if not issue or not resolved_key:
            continue

        row_key = _row_issue_key(ws, row, id_col, repo_col)
        if resolved_key != row_key:
            if resolved_key in resolved_key_rows:
                rows_to_delete.add(row)
                continue
            if repo_col:
                ws.cell(row=row, column=repo_col).value = repo_display_name(
                    get_gitlab_repo(issue)
                )

        if resolved_key in resolved_key_rows:
            other = resolved_key_rows[resolved_key]
            if _row_filled_score(ws, row, max_col) > _row_filled_score(ws, other, max_col):
                rows_to_delete.add(other)
                resolved_key_rows[resolved_key] = row
            else:
                rows_to_delete.add(row)
        else:
            resolved_key_rows[resolved_key] = row

    for row in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(row, 1)

    return len(rows_to_delete)


def _sync_issue_links_and_repos(
    ws: Worksheet,
    data_start_row: int,
    columns: Dict[str, int],
    repo_col: Optional[int],
    issues_by_id: Dict[str, Dict],
) -> int:
    """Aplica hyperlinks na coluna # e rotulo do repositorio em todas as linhas."""
    id_col = columns.get("id")
    title_col = columns.get("title")
    if not id_col:
        return 0

    updated = 0
    for row in range(data_start_row, ws.max_row + 1):
        iid = _normalize_id(ws.cell(row=row, column=id_col).value)
        if not iid:
            continue

        issue, issue_key = _match_issue_for_row(
            ws, row, id_col, repo_col, title_col, issues_by_id
        )
        if issue:
            repo = get_gitlab_repo(issue)
        else:
            _, repo = parse_issue_key(_row_issue_key(ws, row, id_col, repo_col) or "")

        _set_issue_id_cell(ws, row, id_col, iid, repo)
        if repo_col:
            ws.cell(row=row, column=repo_col).value = repo_display_name(repo)
        updated += 1

    return updated


def _write_issue_row(
    ws: Worksheet,
    row_num: int,
    columns: Dict[str, int],
    header_map: Dict[str, int],
    issue: Dict,
    issue_id: str,
    title: str,
    module: str,
    area: str,
    created_date: Optional[datetime],
    fill_metadata: bool,
    *,
    preserve_module: bool = False,
    preserve_area: bool = False,
) -> int:
    """Escreve campos basicos e, se solicitado, metadados do GitLab."""
    written = 0
    repo = get_gitlab_repo(issue)
    if 'id' in columns:
        _set_issue_id_cell(ws, row_num, columns['id'], issue_id, repo)
        written += 1
    if 'title' in columns:
        ws.cell(row=row_num, column=columns['title']).value = title
        written += 1
    if 'module' in columns:
        current_module = ws.cell(row=row_num, column=columns['module']).value
        if not (preserve_module and current_module not in (None, '')):
            ws.cell(row=row_num, column=columns['module']).value = module
        written += 1
    if 'area' in columns:
        current_area = ws.cell(row=row_num, column=columns['area']).value
        if preserve_area and current_area not in (None, ''):
            pass
        elif area or current_area in (None, ''):
            ws.cell(row=row_num, column=columns['area']).value = area or current_area
        written += 1
    if 'created' in columns and created_date:
        ws.cell(row=row_num, column=columns['created']).value = created_date
        written += 1

    if fill_metadata:
        _sync_gitlab_metadata(ws, row_num, header_map, issue, created_date)
        written += 1

    repo_col = header_map.get('repositorio')
    if repo_col:
        ws.cell(row=row_num, column=repo_col).value = repo_display_name(repo)

    return written


def print_final_report(stats: Dict, excel_file: str) -> None:
    """Imprime o resumo de execucao (reutilizavel, evita duplicacao)."""
    print("\n" + "=" * 70)
    print("PROCESSAMENTO CONCLUÍDO")
    print("=" * 70)
    print(f"Total de issues extraídas do JSON:     {stats['total_extracted']}")
    print(f"Issues após data de corte:              {stats['after_cutoff']}")
    print(f"Issues descartadas (antes do corte):    {stats['before_cutoff']}")
    print(f"Issues novas ignoradas (módulo):        {stats['skipped_module']}")
    print(f"\n-> Issues ATUALIZADAS (preserve columns): {stats['updated_existing']}")
    print(f"-> Issues NOVAS (inseridas):              {stats['new_added']}")
    print(f"-> Metadados GitLab preenchidos:          {stats['metadata_filled']}")
    print(f"-> Tipos inferidos (sem label tipo::):      {stats.get('tipos_inferidos', 0)}")
    print(f"-> Issues enriquecidas (Dev/Git):           {stats.get('dev_enriquecidas', 0)}")
    print(f"-> Issues com branch vinculada:              {stats.get('dev_com_branch', 0)}")
    print(f"-> Areas funcionais atualizadas:          {stats['areas_atualizadas']}")
    print(f"\nMódulos encontrados: {len(stats['modules_found'])}")
    for module, count in sorted(stats['modules_found'].items(), key=lambda x: -x[1]):
        print(f"  • {module}: {count} issues")
    print("\nAVISO - COLUNAS MANUAIS (nao alteradas):")
    for col in sorted(MANUAL_ONLY_COLUMNS):
        print(f"  • {col}")
    print(f"\nOK - Arquivo salvo: {excel_file}")
    print("=" * 70)


def _load_issues_from_disk() -> Optional[List[Dict]]:
    """Carrega issues do gitlab_issues_raw.json (procura em multiplos locais)."""
    json_paths = [
        'gitlab_issues_raw.json',
        'mgi/gitlab_issues_raw.json',
        os.path.join(os.path.dirname(__file__), 'gitlab_issues_raw.json'),
    ]
    for json_path in json_paths:
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                issues = json.load(f)
            print(f"OK - Carregado: {json_path}")
            return issues
        except FileNotFoundError:
            continue
    return None


def _unavailable_optional_features() -> List[str]:
    """Lista recursos opcionais que NAO foram carregados (import falhou -> None).

    Como todos os modulos sao locais, um valor None aqui indica um problema de
    import (ex.: dependencia circular ou erro de sintaxe), nunca uma dependencia
    externa ausente. Serve para evitar falhas silenciosas.
    """
    features = {
        "deteccao de Area (detectar_area_funcional.build_detector)": build_detector,
        "save COM (excel_com_save.save_workbook_preserving_filters)": save_workbook_preserving_filters,
        "filtro de fechadas (issue_filters.filtrar_issues_fechadas_antigas)": filtrar_issues_fechadas_antigas,
        "inferencia de Tipo (inferir_tipo_issue.build_tipo_detector)": build_tipo_detector,
        "enriquecimento Dev/Git (enriquecer_dev_git.build_dev_enricher)": build_dev_enricher,
        "KPI Parceria (atualizar_dashboard_kpis.atualizar_kpi_parceria)": atualizar_kpi_parceria,
        "graficos novos (atualizar_graficos_dashboard.atualizar_graficos_novos)": atualizar_graficos_novos,
        "listas taxonomia (atualizar_listas_taxonomia.sync_listas_taxonomia)": sync_listas_taxonomia,
        "qualidade de dados (qualidade_dados.atualizar_qualidade_dados)": atualizar_qualidade_dados,
        "releases dashboard (atualizar_releases_dashboard)": atualizar_releases_dashboard,
        "relatorio de excecoes (relatorio_excecoes.coletar_excecoes_wb)": coletar_excecoes_wb,
        "normalizacao de modulos (modulo_normalization.apply_module_normalization)": apply_module_normalization,
    }
    return [name for name, symbol in features.items() if symbol is None]


def _upsert_issues(
    ws: Worksheet,
    gitlab_issues: List[Dict],
    columns: Dict[str, int],
    header_map: Dict[str, int],
    *,
    existing_ids: Dict[str, int],
    initial_existing_ids: Set[str],
    preserve_taxonomy: bool,
    cutoff_date: datetime,
    full_refresh: bool,
    all_modules: bool,
    area_detector: Optional["AreaFuncionalDetector"],
    tipo_detector: Optional["TipoIssueDetector"],
    stats: Dict,
    active_ids: Set[str],
    inferred_tipo_ids: Set[str],
    processed_issue_ids: Set[str],
) -> None:
    """Aplica upsert das issues na aba Dados (atualiza existentes, insere novas).

    Muta ``ws``, ``stats``, ``existing_ids`` e os conjuntos de ids passados.
    """
    # Process each issue
    for issue in gitlab_issues:
        issue_id = str(issue.get('id', ''))
        issue_key = make_issue_key(issue)
        title = issue.get('title', '')
        created_date_str = issue.get('createdDate', '')

        stats['total_extracted'] += 1

        created_date = parse_date(created_date_str)
        module = extract_module(title)
        area_detection = _resolve_area_detection(issue, title, area_detector)
        area = area_detection.area

        if module:
            stats['modules_found'][module] = stats['modules_found'].get(module, 0) + 1

        if created_date and created_date < cutoff_date:
            stats['before_cutoff'] += 1
            if not (full_refresh and issue_key in existing_ids):
                continue

        # ISSUES JA NA PLANILHA: sincronizar titulo; preservar modulo/area se ja preenchidos
        if issue_key in existing_ids:
            active_ids.add(issue_key)
            row_num = existing_ids[issue_key]
            needs_metadata = full_refresh or _row_needs_gitlab_metadata(ws, row_num, header_map)
            preserve_row = preserve_taxonomy and issue_key in initial_existing_ids
            area_col = columns.get("area")
            area_before = (
                ws.cell(row=row_num, column=area_col).value if area_col else None
            )
            written = _write_issue_row(
                ws,
                row_num,
                columns,
                header_map,
                issue,
                issue_id,
                title,
                module,
                area,
                created_date,
                fill_metadata=needs_metadata,
                preserve_module=preserve_row,
                preserve_area=preserve_row,
            )
            _sync_label_columns(ws, row_num, header_map, issue)
            stats['updated_existing'] += 1
            if needs_metadata:
                stats['metadata_filled'] += 1
            processed_issue_ids.add(issue_key)
            if _sync_tipo_column(
                ws, row_num, header_map, issue, tipo_detector, inferred_tipo_ids
            ):
                stats['tipos_inferidos'] += 1
            if area_col and not preserve_row:
                area_after = ws.cell(row=row_num, column=area_col).value
                if area_after not in (None, "") and area_after != area_before:
                    stats['areas_atualizadas'] += 1
            elif area_col and preserve_row and area_before in (None, "") and area:
                stats['areas_atualizadas'] += 1
            continue

        stats['after_cutoff'] += 1

        # ISSUES NOVAS: aplicar filtro de modulo
        allow_new = _config.ALLOW_NEW_ISSUES if _config else True
        if not allow_new:
            stats['skipped_module'] += 1
            continue
        if not _modulo_permitido(module, all_modules=all_modules):
            stats['skipped_module'] += 1
            continue

        active_ids.add(issue_key)
        new_row = ws.max_row + 1
        written = _write_issue_row(
            ws,
            new_row,
            columns,
            header_map,
            issue,
            issue_id,
            title,
            module,
            area,
            created_date,
            fill_metadata=True,
        )
        if written > 0:
            existing_ids[issue_key] = new_row
            stats['new_added'] += 1
            stats['metadata_filled'] += 1
            processed_issue_ids.add(issue_key)
            if _sync_tipo_column(
                ws, new_row, header_map, issue, tipo_detector, inferred_tipo_ids
            ):
                stats['tipos_inferidos'] += 1
            if area:
                stats['areas_atualizadas'] += 1


def _postprocess_rows(
    ws: Worksheet,
    wb,
    *,
    columns: Dict[str, int],
    header_map: Dict[str, int],
    data_start_row: int,
    id_col: int,
    repo_col: Optional[int],
    issues_by_id: Dict[str, Dict],
    area_detector: Optional["AreaFuncionalDetector"],
    tipo_detector: Optional["TipoIssueDetector"],
    dev_enricher: Optional["GitDevEnricher"],
    inferred_tipo_ids: Set[str],
    processed_issue_ids: Set[str],
    active_ids: Set[str],
    initial_load: bool,
    cutoff_date: datetime,
    stats: Dict,
) -> Tuple[int, int, List[Tuple[int, int, int]]]:
    """Backfill de area/tipo, dedupe/remocao, ordenacao, links, enriquecimento Dev/Git,
    destaques visuais e expansao da tabela.

    Retorna ``(last_data_row, sorted_rows, highlight_cells)`` e muta ``ws``/``wb``/``stats``.
    """
    area_force = os.environ.get("MGI_AREA_FORCE", "0").lower() not in (
        "0", "false", "no",
    )
    areas_backfill = _backfill_areas_for_sheet(
        ws,
        columns,
        data_start_row,
        id_col,
        repo_col,
        issues_by_id,
        area_detector,
        force=area_force,
    )
    if areas_backfill:
        stats["areas_atualizadas"] += areas_backfill
        print(f"OK - {areas_backfill} areas funcionais preenchidas/corrigidas na aba Dados")

    tipos_preenchidos = _fill_missing_tipos_in_sheet(
        ws,
        header_map,
        data_start_row,
        id_col,
        repo_col,
        issues_by_id,
        tipo_detector,
        inferred_tipo_ids,
    )
    if tipos_preenchidos:
        stats['tipos_inferidos'] += tipos_preenchidos
        print(f"OK - {tipos_preenchidos} tipos inferidos em linhas com Tipo vazio")

    if initial_load:
        removed = _remove_rows_not_in_active_set(
            ws, data_start_row, id_col, repo_col, active_ids
        )
        stats['rows_removed'] = removed
        if removed:
            print(
                f"OK - Carga inicial: {removed} linhas removidas "
                f"(issues anteriores a {cutoff_date.strftime('%d/%m/%Y')} ou fora dos criterios)"
            )

    deduped = _dedupe_dados_rows(
        ws,
        data_start_row,
        id_col,
        repo_col,
        columns.get("title"),
        issues_by_id,
    )
    if deduped:
        stats["rows_removed"] = stats.get("rows_removed", 0) + deduped
        print(f"OK - {deduped} linhas duplicadas removidas (chave repositorio:#)")

    sorted_rows = _sort_issue_rows_desc(ws, data_start_row, id_col, repo_col)
    print(f"OK - Issues ordenadas por # decrescente ({sorted_rows} linhas)")

    links = _sync_issue_links_and_repos(
        ws, data_start_row, columns, repo_col, issues_by_id
    )
    if links:
        print(f"OK - {links} hyperlinks GitLab aplicados na coluna #")
    last_data_row = data_start_row + sorted_rows - 1 if sorted_rows else data_start_row - 1

    dev_highlights, dev_enriched, dev_with_branch = _enrich_dev_git_for_sheet(
        ws,
        header_map,
        data_start_row,
        id_col,
        repo_col,
        issues_by_id,
        dev_enricher,
    )
    stats['dev_enriquecidas'] = dev_enriched
    stats['dev_com_branch'] = dev_with_branch
    if dev_enriched:
        print(
            f"OK - {dev_enriched} issues enriquecidas com dados Git "
            f"({dev_with_branch} com branch vinculada)"
        )

    tipo_highlights = _apply_inferred_tipo_highlights(
        ws,
        header_map,
        data_start_row,
        id_col,
        repo_col,
        inferred_tipo_ids,
        processed_issue_ids,
    )
    if tipo_highlights:
        print(f"OK - {len(tipo_highlights)} celulas Tipo destacadas em laranja (inferidas)")

    highlight_cells: List[Tuple[int, int, int]] = []
    highlight_cells.extend(tipo_highlights)
    highlight_cells.extend(dev_highlights)

    # Expandir tabela Excel
    new_table_ref, formulas_applied = _finalize_sheet_after_inserts(ws, data_start_row, header_map)
    if new_table_ref:
        print(f"OK - Tabela {ISSUES_TABLE_NAME} expandida para: {new_table_ref}")
    if formulas_applied:
        print(f"OK - Formulas aplicadas em {formulas_applied} linhas")

    _repair_calc_formulas(wb)

    return last_data_row, sorted_rows, highlight_cells


def _run_dashboard_hooks(
    wb,
    ws: Worksheet,
    *,
    header_row: int,
    data_start_row: int,
    last_data_row: int,
    id_col: int,
    repo_col: Optional[int],
    issues_by_id: Dict[str, Dict],
    area_detector: Optional["AreaFuncionalDetector"],
    full_refresh: bool,
    preserve_taxonomy: bool,
    excel_file: str,
    stats: Dict,
) -> None:
    """Executa os hooks opcionais de dashboard (KPIs, graficos, listas, qualidade,
    releases e relatorio de excecoes). Cada hook so roda se o modulo foi carregado."""
    if atualizar_kpi_parceria:
        kpi_stats = atualizar_kpi_parceria(wb, last_data_row)
        print(
            f"OK - KPI Parceria: {kpi_stats['parcerias']} parcerias, "
            f"grafico e filtro em U4 "
            f"({kpi_stats['formulas_updated']} formulas KPI atualizadas)"
        )

    if atualizar_graficos_novos:
        graf_stats = atualizar_graficos_novos(wb, last_data_row)
        criados = [k for k, v in graf_stats.items() if not v.get("skipped")]
        print(f"OK - Graficos novos: {', '.join(criados)}")

    if sync_listas_taxonomia:
        listas_stats = sync_listas_taxonomia(wb)
        print(
            f"OK - Listas taxonomia: {listas_stats['modulos_canonicos']} canonicos, "
            f"{listas_stats['de_para']} de-para, "
            f"{listas_stats['areas_padrao']} areas padrao"
        )

    if ensure_module_cols and apply_module_normalization:
        mod_cols = ensure_module_cols(ws, header_row)
        mod_stats = apply_module_normalization(
            ws,
            header_row,
            data_start_row,
            last_data_row,
            mod_cols,
            sync_modulo_column=True,
            preserve_filled_module=preserve_taxonomy,
        )
        stats["modulos_canonicos"] = mod_stats.get("canonicos", 0)
        stats["modulos_custom"] = mod_stats.get("custom", 0)
        preserved = mod_stats.get("preservados", 0)
        print(
            f"OK - Modulos normalizados: {mod_stats['canonicos']} canonicos, "
            f"{mod_stats['custom']} custom, {mod_stats['vazios']} vazios"
            + (f", {preserved} linhas com Módulo preservado" if preserved else "")
        )

    if atualizar_qualidade_dados:
        qstats = atualizar_qualidade_dados(wb, header_row, data_start_row, last_data_row)
        stats["qualidade_formulas"] = qstats["formulas_aplicadas"]
        print(
            f"OK - Qualidade dos dados: {qstats['formulas_aplicadas']} formulas, "
            f"{qstats['calc_metricas']} metricas em _Calc, "
            f"grafico categoria ({qstats.get('categoria_categorias', 0)} grupos)"
        )

    if ensure_quality_columns and write_confidence:
        quality_cols = ensure_quality_columns(ws, header_row)
        conf_col = quality_cols["confianca area"]
        conf_written = 0
        for row in range(data_start_row, last_data_row + 1):
            issue_key = _row_issue_key(ws, row, id_col, repo_col)
            if not issue_key:
                continue
            issue = lookup_issue(issues_by_id, issue_key)
            if not issue:
                continue
            detection = _resolve_area_detection(
                issue, issue.get("title", ""), area_detector
            )
            if detection.confidence > 0:
                write_confidence(
                    ws,
                    row,
                    conf_col,
                    detection.confidence,
                    overwrite=full_refresh,
                )
                conf_written += 1
        if conf_written:
            print(f"OK - Confianca Area (Git) em {conf_written} linhas")

    if atualizar_releases_dashboard:
        rstats = atualizar_releases_dashboard(wb)
        stats["releases_total"] = rstats.get("releases_total", 0)
        if rstats.get("fonte_ok"):
            print(
                f"OK - Releases Git: {rstats['releases_total']} tags, "
                f"{rstats['releases_exibidas']} exibidas em _Calc"
            )
        else:
            print("AVISO - gitlab_git_data.json ausente; grafico de releases omitido")

    if coletar_excecoes_wb and exportar_excecoes:
        # Bloco opcional de relatorio: nunca pode abortar a gravacao do Excel.
        try:
            excecoes = coletar_excecoes_wb(wb)
            logs_dir = Path(_config.LOGS_DIR) if _config else Path(excel_file).parent / "logs"
            exc_paths = exportar_excecoes(excecoes, logs_dir)
            stats["excecoes_qualidade"] = len(excecoes)
            print(f"OK - {len(excecoes)} excecoes -> {exc_paths['csv']}")
        except Exception as exc:
            print(f"AVISO - relatorio de excecoes ignorado: {exc}")


def _save_dashboard(
    wb,
    ws: Worksheet,
    excel_file: str,
    *,
    header_row: int,
    data_start_row: int,
    last_data_row: int,
    sorted_rows: int,
    highlight_cells: List[Tuple[int, int, int]],
) -> Optional[str]:
    """Grava o workbook (openpyxl + COM para preservar validacoes/filtros).

    Retorna o caminho final salvo (pode ser uma copia *_atualizado.xlsx se o
    arquivo estava aberto) ou ``None`` quando nao foi possivel gravar.
    """
    _repair_calc_formulas(wb)
    save_path = excel_file
    openpyxl_saved = False
    try:
        wb.save(excel_file)
        openpyxl_saved = True
    except PermissionError:
        alt_path = str(
            Path(excel_file).with_name(
                f"{Path(excel_file).stem}_atualizado{Path(excel_file).suffix}"
            )
        )
        print(f"\nAVISO - {excel_file} em uso (feche o Excel para sobrescrever).")
        print(f"        Gravando copia em: {alt_path}")
        try:
            wb.save(alt_path)
            save_path = alt_path
            openpyxl_saved = True
        except PermissionError:
            print("\nERRO - Nao foi possivel gravar o Excel.")
            print(f"       Feche o arquivo antes de executar: {excel_file}")
            return None
    except OSError as exc:
        print(f"\nERRO - Falha ao gravar o Excel: {exc}")
        return None

    if save_path != excel_file:
        excel_file = save_path

    # Save the workbook (prefer COM to preserve Listas / data validation on Dados)
    saved = False
    if save_workbook_preserving_filters and sorted_rows > 0:
        saved = save_workbook_preserving_filters(
            excel_file,
            ws,
            header_row,
            data_start_row,
            last_data_row=last_data_row,
            expected_data_rows=sorted_rows,
            highlight_cells=highlight_cells,
        )

    if not saved:
        print(
            "AVISO - Gravacao via openpyxl pode remover validacoes de dados "
            "e afetar filtros. Instale pywin32 e use Excel no Windows para preservar."
        )
        try:
            wb.save(excel_file)
            saved = True
        except PermissionError:
            if openpyxl_saved:
                saved = True
                print(
                    "\nAVISO - Excel aberto: gravacao COM/openpyxl adicional ignorada; "
                    f"arquivo ja salvo em {save_path}"
                )
            else:
                print("\nERRO - Nao foi possivel gravar o Excel.")
                print(f"       Feche o arquivo antes de executar: {excel_file}")
                return None
        except OSError as exc:
            print(f"\nERRO - Falha ao gravar o Excel: {exc}")
            return None

    if not saved:
        print("\nERRO - Nao foi possivel gravar o Excel.")
        print(f"       Feche o MGI_Dashboard.xlsx e execute novamente.")
        return None

    return excel_file


def _load_and_index_issues(
    issues: Optional[List[Dict]],
    full_refresh: bool,
) -> Optional[Tuple[List[Dict], Dict[str, Dict]]]:
    """Carrega as issues (memoria ou disco), indexa por chave e aplica o filtro de
    fechadas antigas. Retorna ``(gitlab_issues, issues_by_id)`` ou ``None`` sem fonte."""
    gitlab_issues = issues if issues is not None else _load_issues_from_disk()

    if gitlab_issues is None:
        print("ERRO - gitlab_issues_raw.json not found!")
        return None

    raw_issues = list(gitlab_issues)
    repo_counts, missing_repo = summarize_issues_by_repo(raw_issues)
    print(f"OK - Issues por repositorio no JSON: {repo_counts}")
    if missing_repo:
        print(
            f"AVISO - {missing_repo} issues sem gitlab_repo no JSON "
            f"(serao tratadas como {DEFAULT_GITLAB_REPO})"
        )

    issues_by_id: Dict[str, Dict] = {}
    for issue in raw_issues:
        iid = str(issue.get("id", "")).strip()
        if not iid:
            continue
        key = make_issue_key(issue)
        issues_by_id[key] = issue

    if filtrar_issues_fechadas_antigas:
        exclude_days = (
            _config.closed_exclude_days()
            if _config and hasattr(_config, "closed_exclude_days")
            else (_config.CLOSED_EXCLUDE_DAYS if _config else 60)
        )
        if _config and getattr(_config, "INITIAL_LOAD", False):
            print("OK - Carga inicial: filtro de issues fechadas DESATIVADO (todas incluidas)")
        elif full_refresh:
            print(
                "OK - Execucao completa: JSON sem filtro de 60 dias para atualizar linhas existentes"
            )
            exclude_days = 0
        gitlab_issues, excluidas = filtrar_issues_fechadas_antigas(
            gitlab_issues, days=exclude_days
        )
        if excluidas:
            print(
                f"OK - {excluidas} issues fechadas ha mais de {exclude_days} dias "
                f"ignoradas ({len(gitlab_issues)} a processar)"
            )
        elif exclude_days <= 0:
            print(f"OK - {len(gitlab_issues)} issues a processar (sem filtro de fechadas)")

    return gitlab_issues, issues_by_id


def _build_runtime_detectors() -> Tuple[
    Optional["AreaFuncionalDetector"],
    Optional["TipoIssueDetector"],
    Optional["GitDevEnricher"],
]:
    """Constroi os detectores Git (area, tipo, dev). Em modo rapido
    (MGI_FAST_REPO_SYNC) desliga todos para evitar acesso a repositorios."""
    area_detector = build_detector() if build_detector else None
    if area_detector and os.environ.get("MGI_FAST_REPO_SYNC", "0").lower() not in ("0", "false", "no"):
        area_detector = None
        print("OK - Modo rapido: deteccao de Area Funcional desativada")
    elif area_detector:
        print("OK - Deteccao de Area Funcional via Git habilitada")

    tipo_detector = build_tipo_detector() if build_tipo_detector else None
    if tipo_detector and os.environ.get("MGI_FAST_REPO_SYNC", "0").lower() not in ("0", "false", "no"):
        tipo_detector = None
        print("OK - Modo rapido: inferencia de Tipo desativada")
    elif tipo_detector:
        print("OK - Inferencia de Tipo via Git/titulo habilitada (destaque laranja no Excel)")

    dev_enricher = build_dev_enricher() if build_dev_enricher else None
    if dev_enricher and os.environ.get("MGI_FAST_REPO_SYNC", "0").lower() not in ("0", "false", "no"):
        dev_enricher = None
        print("OK - Modo rapido: enriquecimento Dev/Git desativado")
    elif dev_enricher:
        print("OK - Enriquecimento Dev/Git habilitado (destaque azul claro no Excel)")

    return area_detector, tipo_detector, dev_enricher


def process_issues(
    cutoff_date: Optional[datetime] = None,
    issues: Optional[List[Dict]] = None,
    excel_file: Optional[str] = None,
    all_modules: Optional[bool] = None,
    initial_load: Optional[bool] = None,
    full_refresh: Optional[bool] = None,
) -> Optional[Dict]:
    """Processa issues e atualiza a planilha Excel.

    Args:
        cutoff_date: data minima para filtrar issues (default: 2024-01-01).
        issues: lista de issues ja carregada em memoria. Se None, le do disco.
        excel_file: caminho do Excel de destino (default: MGI_Dashboard.xlsx).
        all_modules: True inclui todos os modulos; None usa config/env.
        initial_load: True ignora apenas o filtro de fechadas antigas (60 dias).
        full_refresh: True reprocessa metadados GitLab e enriquecimentos em todas
            as linhas existentes (ignora filtro de 60 dias para atualizacoes).

    Returns:
        Dicionario de estatisticas, ou None em caso de falha de carregamento.
    """

    # Default cutoff date
    if cutoff_date is None:
        cutoff_date = datetime(2024, 1, 1)

    if all_modules is None:
        all_modules = bool(_config and getattr(_config, 'ALL_MODULES', False))
    if initial_load is None:
        initial_load = bool(_config and getattr(_config, 'INITIAL_LOAD', False))
    if full_refresh is None:
        full_refresh = bool(
            _config and hasattr(_config, "is_full_refresh") and _config.is_full_refresh()
        )

    indisponiveis = _unavailable_optional_features()
    if indisponiveis:
        print("AVISO - Recursos opcionais NAO carregados (possivel erro de import/ciclo):")
        for nome in indisponiveis:
            print(f"        - {nome}")

    # Load existing spreadsheet
    if excel_file is None:
        excel_file = (
            str(_config.EXCEL_OUTPUT)
            if _config is not None and hasattr(_config, "EXCEL_OUTPUT")
            else 'MGI_Dashboard.xlsx'
        )
    wb = load_workbook(excel_file)
    ws = wb['Dados']  # Somente a aba Dados e alterada; Listas/filtros nao sao tocados
    protected = ", ".join(
        name.encode("ascii", "replace").decode("ascii")
        for name in sorted(PROTECTED_SHEETS)
    )
    print(f"OK - Abas protegidas (nao editadas): {protected}")

    header_row, data_start_row, columns = _resolve_sheet_layout(ws)
    header_map = _build_full_header_map(ws, header_row)
    header_map = _ensure_repositorio_column(ws, header_row, header_map)
    header_map = _ensure_extra_label_columns(ws, header_row, header_map)
    header_map = _ensure_dev_git_columns(ws, header_row, header_map)
    repo_col = header_map.get('repositorio')

    # Cabecalhos legiveis para log
    header_names = [
        ws.cell(row=header_row, column=c).value
        for c in range(1, ws.max_column + 1)
        if ws.cell(row=header_row, column=c).value
    ]

    print(f"OK - Cabecalho na linha {header_row}, dados a partir da linha {data_start_row}")
    print(f"OK - Colunas mapeadas: {columns}")
    print(f"OK - Cabecalhos: {header_names[:8]}{'...' if len(header_names) > 8 else ''}")

    if 'id' not in columns:
        print("ERRO - coluna de ID nao encontrada na planilha (esperado '#' ou 'ID')")
        return None

    # Get existing IDs and their row mapping (chave composta repositorio:iid)
    existing_ids: Dict[str, int] = {}
    id_col = columns['id']
    for row in range(data_start_row, ws.max_row + 1):
        issue_key = _row_issue_key(ws, row, id_col, repo_col)
        if issue_key:
            existing_ids[issue_key] = row

    print(f"OK - Found {len(existing_ids)} existing IDs in spreadsheet")
    initial_existing_ids: Set[str] = set(existing_ids.keys())
    preserve_taxonomy = (
        _config.PRESERVE_EXISTING_TAXONOMY
        if _config is not None and hasattr(_config, "PRESERVE_EXISTING_TAXONOMY")
        else os.environ.get("MGI_PRESERVE_TAXONOMY", "1").lower() not in ("0", "false", "no")
    )
    if preserve_taxonomy:
        print(
            "OK - Taxonomia preservada: Módulo e Área Funcional nao serao "
            "sobrescritos em issues ja na planilha"
        )

    # Date filter
    print(f"OK - Filtro de data: >= {cutoff_date.strftime('%d/%m/%Y')} (sempre aplicado)")
    if initial_load:
        print("OK - Carga inicial: apenas o filtro de 60 dias fechadas fica desativado")
    if full_refresh:
        print("OK - Execucao COMPLETA: reprocessa metadados, labels, tipo e Dev/Git")
        if preserve_taxonomy:
            print("OK - Execucao completa: Módulo/Área preservados em linhas existentes")
    print(f"OK - Filtro de modulo: {_descricao_filtro_modulos(all_modules)}")

    # Issues em memoria (vindas do maestro) ou carregadas do disco
    loaded = _load_and_index_issues(issues, full_refresh)
    if loaded is None:
        return None
    gitlab_issues, issues_by_id = loaded

    area_detector, tipo_detector, dev_enricher = _build_runtime_detectors()

    # Statistics
    stats = {
        'total_extracted': 0,
        'after_cutoff': 0,
        'before_cutoff': 0,
        'skipped_module': 0,
        'updated_existing': 0,
        'new_added': 0,
        'metadata_filled': 0,
        'tipos_inferidos': 0,
        'dev_enriquecidas': 0,
        'dev_com_branch': 0,
        'areas_atualizadas': 0,
        'rows_removed': 0,
        'modules_found': {},
    }

    active_ids: Set[str] = set()
    inferred_tipo_ids: Set[str] = set()
    processed_issue_ids: Set[str] = set()

    _upsert_issues(
        ws,
        gitlab_issues,
        columns,
        header_map,
        existing_ids=existing_ids,
        initial_existing_ids=initial_existing_ids,
        preserve_taxonomy=preserve_taxonomy,
        cutoff_date=cutoff_date,
        full_refresh=full_refresh,
        all_modules=all_modules,
        area_detector=area_detector,
        tipo_detector=tipo_detector,
        stats=stats,
        active_ids=active_ids,
        inferred_tipo_ids=inferred_tipo_ids,
        processed_issue_ids=processed_issue_ids,
    )

    last_data_row, sorted_rows, highlight_cells = _postprocess_rows(
        ws,
        wb,
        columns=columns,
        header_map=header_map,
        data_start_row=data_start_row,
        id_col=id_col,
        repo_col=repo_col,
        issues_by_id=issues_by_id,
        area_detector=area_detector,
        tipo_detector=tipo_detector,
        dev_enricher=dev_enricher,
        inferred_tipo_ids=inferred_tipo_ids,
        processed_issue_ids=processed_issue_ids,
        active_ids=active_ids,
        initial_load=initial_load,
        cutoff_date=cutoff_date,
        stats=stats,
    )

    _run_dashboard_hooks(
        wb,
        ws,
        header_row=header_row,
        data_start_row=data_start_row,
        last_data_row=last_data_row,
        id_col=id_col,
        repo_col=repo_col,
        issues_by_id=issues_by_id,
        area_detector=area_detector,
        full_refresh=full_refresh,
        preserve_taxonomy=preserve_taxonomy,
        excel_file=excel_file,
        stats=stats,
    )

    saved_path = _save_dashboard(
        wb,
        ws,
        excel_file,
        header_row=header_row,
        data_start_row=data_start_row,
        last_data_row=last_data_row,
        sorted_rows=sorted_rows,
        highlight_cells=highlight_cells,
    )
    if saved_path is None:
        return None
    excel_file = saved_path

    # Print results
    print_final_report(stats, excel_file)

    return stats

if __name__ == '__main__':
    process_issues()
