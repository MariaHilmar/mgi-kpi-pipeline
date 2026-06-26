#!/usr/bin/env python3
"""
Normaliza modulos na planilha (12 canonicos) e exporta sugestoes de titulo GitLab.

Uso:
  python normalizar_modulos_planilha.py
  python normalizar_modulos_planilha.py --excel D:\\MGI-Relatórios\\MGI_Dashboard.xlsx
  python normalizar_modulos_planilha.py --formulas   # usa VLOOKUP em vez de valores
"""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

try:
    import config
except ImportError:
    config = None

from atualizar_listas_taxonomia import sync_listas_taxonomia
from modulo_normalization import (
    apply_module_formulas,
    apply_module_normalization,
    ensure_module_columns,
)
from process_gitlab_issues_v2 import _repair_calc_formulas, _resolve_sheet_layout
from qualidade_dados import atualizar_qualidade_dados
from taxonomy import CUSTOM_BUCKET, suggest_title_module_fix


def exportar_sugestoes_titulo(
    ws,
    header_row: int,
    data_start: int,
    last_row: int,
    output_dir: Path,
) -> Path:
    import unicodedata

    def norm(s: str) -> str:
        t = unicodedata.normalize("NFKD", (s or "").strip())
        return "".join(c for c in t if not unicodedata.combining(c)).casefold()

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v:
            headers[norm(str(v))] = c

    id_col = headers.get("#") or 1
    title_col = headers.get("titulo")
    repo_col = headers.get("repositorio")
    orig_col = headers.get("modulo original")
    norm_col = headers.get("modulo normalizado")

    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = output_dir / f"sugestoes_titulo_gitlab_{stamp}.csv"

    rows = []
    for row in range(data_start, last_row + 1):
        iid = ws.cell(row, id_col).value
        if iid in (None, "", "#"):
            continue
        title = str(ws.cell(row, title_col).value or "") if title_col else ""
        sugestao = suggest_title_module_fix(title)
        if not sugestao:
            continue
        rows.append(
            {
                "repositorio": str(ws.cell(row, repo_col).value or "") if repo_col else "",
                "issue": str(iid),
                "titulo_atual": title[:250],
                "titulo_sugerido": sugestao[:250],
                "modulo_original": str(ws.cell(row, orig_col).value or "") if orig_col else "",
                "modulo_normalizado": str(ws.cell(row, norm_col).value or "") if norm_col else "",
            }
        )

    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "repositorio",
                "issue",
                "titulo_atual",
                "titulo_sugerido",
                "modulo_original",
                "modulo_normalizado",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    return path


def main() -> int:
    default_excel = config.EXCEL_OUTPUT if config else Path(r"D:\MGI-Relatórios\MGI_Dashboard.xlsx")
    default_logs = config.LOGS_DIR if config else Path(r"D:\MGI-Relatórios\logs")

    parser = argparse.ArgumentParser(description="Normaliza modulos para 12 canonicos")
    parser.add_argument("--excel", type=Path, default=default_excel)
    parser.add_argument("--output-dir", type=Path, default=default_logs)
    parser.add_argument(
        "--formulas",
        action="store_true",
        help="Grava formulas VLOOKUP (padrao: valores calculados)",
    )
    parser.add_argument(
        "--resync-all",
        action="store_true",
        help="Reprocessa Módulo/Original/Normalizado mesmo quando Módulo ja esta preenchido",
    )
    parser.add_argument(
        "--sem-sync-modulo",
        action="store_true",
        help="Nao altera coluna Módulo existente",
    )
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"ERRO - Excel nao encontrado: {args.excel}")
        return 1

    from openpyxl import load_workbook

    print(f"OK - Carregando {args.excel}")
    wb = load_workbook(args.excel)
    ws = wb["Dados"]
    header_row, data_start, _ = _resolve_sheet_layout(ws)
    last_row = ws.max_row
    while last_row > data_start - 1 and ws.cell(last_row, 1).value in (None, ""):
        last_row -= 1

    listas = sync_listas_taxonomia(wb)
    print(
        f"OK - Listas: {listas['modulos_canonicos']} canonicos, "
        f"{listas['de_para']} entradas de-para"
    )

    module_cols = ensure_module_columns(ws, header_row)
    if args.formulas:
        n = apply_module_formulas(ws, header_row, data_start, last_row, module_cols)
        print(f"OK - Formulas modulo em {n} linhas")
    else:
        stats = apply_module_normalization(
            ws,
            header_row,
            data_start,
            last_row,
            module_cols,
            sync_modulo_column=not args.sem_sync_modulo,
            preserve_filled_module=not args.resync_all,
        )
        preserved = stats.get("preservados", 0)
        print(
            f"OK - Normalizacao: {stats['canonicos']} canonicos, "
            f"{stats['custom']} custom, {stats['vazios']} vazios, "
            f"{stats['modulo_sincronizado']} coluna Módulo atualizada"
            + (f", {preserved} linhas preservadas" if preserved else "")
        )

    qualidade = atualizar_qualidade_dados(wb, header_row, data_start, last_row)
    print(f"OK - Qualidade: {qualidade['formulas_aplicadas']} formulas KPI")

    sugestoes = exportar_sugestoes_titulo(ws, header_row, data_start, last_row, args.output_dir)
    print(f"OK - Sugestoes titulo GitLab: {sugestoes}")

    _repair_calc_formulas(wb)

    try:
        wb.save(args.excel)
        print(f"OK - Salvo: {args.excel}")
    except PermissionError:
        alt = args.excel.with_name(f"{args.excel.stem}_modulos{args.excel.suffix}")
        wb.save(alt)
        print(f"AVISO - Arquivo em uso. Salvo em: {alt}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
