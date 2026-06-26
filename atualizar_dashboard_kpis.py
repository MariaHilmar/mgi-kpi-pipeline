#!/usr/bin/env python3
"""
Atualiza KPIs e graficos do Dashboard Executivo apos processamento de issues.
"""

from __future__ import annotations

from typing import List, Optional, Sequence, Set, Tuple

from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from calc_formulas import count_by_label_formula

DASHBOARD_SHEET = "Dashboard Executivo"
CALC_SHEET = "_Calc"
LISTAS_SHEET = "Listas"
DADOS_SHEET = "Dados"

PARCERIA_COL_DADOS = 10  # J
ANO_CRIACAO_COL_DADOS = 23  # W

PARCERIA_LABEL_CELL = "T4"
PARCERIA_FILTER_CELL = "U4"
PARCERIA_FILTER_REF = "$U$4"

PARCERIA_CALC_LABEL_COL = 22  # V
PARCERIA_CALC_QTDE_COL = 23  # W
PARCERIA_CALC_START_ROW = 3

PARCERIA_FILTER_TERM = (
    f'*--((({PARCERIA_FILTER_REF}="Todos")+'
    f'(({PARCERIA_FILTER_REF}="Não informado")*(tbIssues[Parceria]=""))+'
    f'(tbIssues[Parceria]={PARCERIA_FILTER_REF}))>0)*'
)

KPI_FORMULA_CELLS: Tuple[str, ...] = (
    "B10", "E10", "H10", "K10", "N10",
    "B12", "E12", "H12", "K12", "N12",
)

SPRINT_FILTER_MARKER = '*--((($R$4="Todos")+(tbIssues[Sprint]=$R$4))>0)*'


def _unique_parcerias(ws_dados: Worksheet, last_row: int) -> List[str]:
    values: Set[str] = set()
    empty = False
    for row in range(3, last_row + 1):
        raw = ws_dados.cell(row=row, column=PARCERIA_COL_DADOS).value
        if raw is None or str(raw).strip() == "":
            empty = True
        else:
            values.add(str(raw).strip())
    ordered = sorted(values)
    if empty:
        ordered.append("Não informado")
    return ordered


def _sync_lista_parcerias(wb: Workbook, parcerias: Sequence[str]) -> None:
    ws = wb[LISTAS_SHEET]
    ws.cell(row=1, column=9).value = "Parceria"
    ws.cell(row=2, column=9).value = "Todos"
    for idx, name in enumerate(parcerias, start=3):
        ws.cell(row=idx, column=9).value = name

    last = 2 + len(parcerias)
    ref = f"{LISTAS_SHEET}!$I$2:$I${last}"
    if "Lista_Parceria" in wb.defined_names:
        del wb.defined_names["Lista_Parceria"]
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names.add(DefinedName("Lista_Parceria", attr_text=ref))


def _sync_calc_parceria(wb: Workbook, parcerias: Sequence[str], last_row: int) -> int:
    ws = wb[CALC_SHEET]
    ws.cell(row=1, column=PARCERIA_CALC_LABEL_COL).value = "Parceria"
    ws.cell(row=2, column=PARCERIA_CALC_LABEL_COL).value = "Parceria"
    ws.cell(row=2, column=PARCERIA_CALC_QTDE_COL).value = "Qtde"

    label_col = get_column_letter(PARCERIA_CALC_LABEL_COL)
    for offset, parceria in enumerate(parcerias):
        row = PARCERIA_CALC_START_ROW + offset
        label_ref = f"{label_col}{row}"
        ws.cell(row=row, column=PARCERIA_CALC_LABEL_COL).value = parceria
        ws.cell(row=row, column=PARCERIA_CALC_QTDE_COL).value = count_by_label_formula(
            label_ref, PARCERIA_COL_DADOS, last_row
        )

    for row in range(PARCERIA_CALC_START_ROW + len(parcerias), ws.max_row + 1):
        if ws.cell(row=row, column=PARCERIA_CALC_LABEL_COL).value:
            ws.cell(row=row, column=PARCERIA_CALC_LABEL_COL).value = None
            ws.cell(row=row, column=PARCERIA_CALC_QTDE_COL).value = None

    return len(parcerias)


def _inject_parceria_filter(formula: str) -> str:
    if not formula or not str(formula).startswith("="):
        return formula
    if "tbIssues[Parceria]" in formula:
        return formula
    if SPRINT_FILTER_MARKER in formula:
        return formula.replace(SPRINT_FILTER_MARKER, SPRINT_FILTER_MARKER + PARCERIA_FILTER_TERM, 1)
    marker = "*(tbIssues[Ano Criação]>=2024)*"
    if marker in formula:
        return formula.replace(marker, PARCERIA_FILTER_TERM + marker, 1)
    return formula


def _sync_dashboard_filters(wb: Workbook) -> None:
    ws = wb[DASHBOARD_SHEET]
    if ws[PARCERIA_FILTER_CELL].value in (None, ""):
        ws[PARCERIA_FILTER_CELL].value = "Todos"
    ws[PARCERIA_LABEL_CELL].value = "Parceria"

    from openpyxl.worksheet.datavalidation import DataValidation

    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation
        if PARCERIA_FILTER_CELL not in str(dv.sqref)
    ]
    dv = DataValidation(type="list", formula1="=Lista_Parceria", allow_blank=False)
    dv.add(ws[PARCERIA_FILTER_CELL])
    ws.add_data_validation(dv)


def _sync_kpi_formulas(wb: Workbook) -> int:
    ws = wb[DASHBOARD_SHEET]
    updated = 0
    for addr in KPI_FORMULA_CELLS:
        cell = ws[addr]
        if not cell.value or not str(cell.value).startswith("="):
            continue
        new_formula = _inject_parceria_filter(str(cell.value))
        if new_formula != cell.value:
            cell.value = new_formula
            updated += 1
    return updated


def _remove_parceria_charts(ws: Worksheet) -> None:
    ws._charts = [ch for ch in ws._charts if not _chart_is_parceria(ch)]


def _chart_is_parceria(chart) -> bool:
    try:
        title = chart.title.tx.rich.p[0].r[0].t
        return title in ("Parcerias", "Parceria")
    except (AttributeError, IndexError, TypeError):
        return False


def _add_parceria_chart(wb: Workbook, num_rows: int) -> None:
    if num_rows <= 0:
        return

    ws_dash = wb[DASHBOARD_SHEET]
    ws_calc = wb[CALC_SHEET]
    _remove_parceria_charts(ws_dash)

    end_row = PARCERIA_CALC_START_ROW + num_rows - 1
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Parcerias"
    chart.y_axis.title = "Issues"
    chart.x_axis.title = "Parceria"

    data = Reference(
        ws_calc,
        min_col=PARCERIA_CALC_QTDE_COL,
        min_row=2,
        max_row=end_row,
    )
    cats = Reference(
        ws_calc,
        min_col=PARCERIA_CALC_LABEL_COL,
        min_row=PARCERIA_CALC_START_ROW,
        max_row=end_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 18

    ws_dash.cell(row=48, column=2).value = "Volume por parceria"
    ws_dash.add_chart(chart, "B49")


def atualizar_kpi_parceria(wb: Workbook, last_data_row: int) -> dict:
    """Configura filtro, formulas, _Calc e grafico de Parceria."""
    ws_dados = wb[DADOS_SHEET]
    parcerias = _unique_parcerias(ws_dados, last_data_row)

    _sync_lista_parcerias(wb, parcerias)
    calc_rows = _sync_calc_parceria(wb, parcerias, last_data_row)
    _sync_dashboard_filters(wb)
    formulas_updated = _sync_kpi_formulas(wb)
    _add_parceria_chart(wb, calc_rows)

    return {
        "parcerias": len(parcerias),
        "calc_rows": calc_rows,
        "formulas_updated": formulas_updated,
    }
