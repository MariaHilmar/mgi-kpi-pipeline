#!/usr/bin/env python3
"""
Gera graficos do Dashboard Executivo com base nos novos dados (Area, Dev, Repo, Git).

Complementa atualizar_dashboard_kpis.py (Parceria).
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass
from typing import List, Optional, Sequence, Tuple

from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from calc_formulas import EMPTY_LABEL, OUTROS_LABEL, count_by_label_formula, outros_count_formula
from process_gitlab_issues_v2 import (
    _build_full_header_map,
    _ensure_dev_git_columns,
    _ensure_repositorio_column,
    _resolve_sheet_layout,
)

DASHBOARD_SHEET = "Dashboard Executivo"
CALC_SHEET = "_Calc"
LISTAS_SHEET = "Listas"
DADOS_SHEET = "Dados"
ANO_CRIACAO_COL = 23  # W


@dataclass
class DimensaoGrafico:
    key: str
    titulo: str
    header_aliases: Tuple[str, ...]
    listas_col: int
    defined_name: str
    calc_label_col: int
    calc_qtde_col: int
    anchor: str
    section_row: int
    section_label: str
    top_n: Optional[int] = None
    horizontal: bool = False


DIMENSOES: Tuple[DimensaoGrafico, ...] = (
    DimensaoGrafico(
        key="repositorio",
        titulo="Issues por Repositório",
        header_aliases=("repositorio", "repositório"),
        listas_col=14,
        defined_name="Lista_Repositorio",
        calc_label_col=34,
        calc_qtde_col=35,
        anchor="J49",
        section_row=48,
        section_label="Volume por repositório",
    ),
    DimensaoGrafico(
        key="area",
        titulo="Área Funcional",
        header_aliases=("área funcional", "area funcional"),
        listas_col=3,
        defined_name="Lista_Área_Funcional",
        calc_label_col=36,
        calc_qtde_col=37,
        anchor="B66",
        section_row=65,
        section_label="Volume por área funcional",
        top_n=14,
    ),
    DimensaoGrafico(
        key="desenvolvedor",
        titulo="Top Desenvolvedores",
        header_aliases=("desenvolvedor",),
        listas_col=15,
        defined_name="Lista_Desenvolvedor",
        calc_label_col=38,
        calc_qtde_col=39,
        anchor="J66",
        section_row=65,
        section_label="Top desenvolvedores (Git)",
        top_n=12,
        horizontal=True,
    ),
    DimensaoGrafico(
        key="dev_mergeado",
        titulo="Merge em master",
        header_aliases=("dev: mergeado?",),
        listas_col=16,
        defined_name="Lista_Dev_Mergeado",
        calc_label_col=40,
        calc_qtde_col=41,
        anchor="B83",
        section_row=82,
        section_label="Issues mergeadas (Git)",
    ),
)


def _resolve_col(header_map: dict, aliases: Sequence[str]) -> Optional[int]:
    for alias in aliases:
        col = header_map.get(alias)
        if col:
            return col
    return None


def _count_by_column(
    ws: Worksheet,
    col: int,
    last_row: int,
    top_n: Optional[int] = None,
    empty_label: str = "Não informado",
) -> List[Tuple[str, int]]:
    counter: Counter = Counter()
    for row in range(3, last_row + 1):
        raw = ws.cell(row=row, column=col).value
        if raw is None or str(raw).strip() == "":
            counter[empty_label] += 1
        else:
            counter[str(raw).strip()] += 1

    ranked = counter.most_common()
    if top_n and len(ranked) > top_n:
        top = ranked[:top_n]
        others = sum(qty for _, qty in ranked[top_n:])
        if others:
            top.append(("Outros", others))
        return top
    return ranked


def _sync_lista(wb: Workbook, dim: DimensaoGrafico, labels: Sequence[str]) -> None:
    ws = wb[LISTAS_SHEET]
    ws.cell(row=1, column=dim.listas_col).value = dim.titulo.split("(")[0].strip()
    ws.cell(row=2, column=dim.listas_col).value = "Todos"
    for idx, name in enumerate(labels, start=3):
        ws.cell(row=idx, column=dim.listas_col).value = name

    last = 2 + len(labels)
    col = get_column_letter(dim.listas_col)
    ref = f"{LISTAS_SHEET}!${col}$2:${col}${last}"
    if dim.defined_name in wb.defined_names:
        del wb.defined_names[dim.defined_name]
    from openpyxl.workbook.defined_name import DefinedName

    wb.defined_names.add(DefinedName(dim.defined_name, attr_text=ref))


def _sync_calc_block(
    ws_calc: Worksheet,
    dim: DimensaoGrafico,
    ranked: Sequence[Tuple[str, int]],
    data_col: int,
    last_row: int,
) -> int:
    label_letter = get_column_letter(dim.calc_label_col)
    qtde_letter = get_column_letter(dim.calc_qtde_col)

    ws_calc.cell(row=1, column=dim.calc_label_col).value = dim.key
    ws_calc.cell(row=2, column=dim.calc_label_col).value = "Rotulo"
    ws_calc.cell(row=2, column=dim.calc_qtde_col).value = "Qtde"

    for offset, (label, _qty) in enumerate(ranked):
        row = 3 + offset
        label_ref = f"{label_letter}{row}"
        ws_calc.cell(row=row, column=dim.calc_label_col).value = label
        if label == OUTROS_LABEL:
            ws_calc.cell(row=row, column=dim.calc_qtde_col).value = outros_count_formula(
                qtde_letter, row, last_row
            )
        else:
            ws_calc.cell(row=row, column=dim.calc_qtde_col).value = count_by_label_formula(
                label_ref, data_col, last_row, empty_label=EMPTY_LABEL
            )

    for row in range(3 + len(ranked), ws_calc.max_row + 1):
        if ws_calc.cell(row=row, column=dim.calc_label_col).value is not None:
            ws_calc.cell(row=row, column=dim.calc_label_col).value = None
            ws_calc.cell(row=row, column=dim.calc_qtde_col).value = None

    return len(ranked)


def _remove_charts_by_title(ws: Worksheet, titles: Sequence[str]) -> None:
    titles_set = set(titles)
    ws._charts = [
        ch
        for ch in ws._charts
        if not _chart_title(ch) in titles_set
    ]


def _chart_title(chart) -> str:
    try:
        return chart.title.tx.rich.p[0].r[0].t
    except (AttributeError, IndexError, TypeError):
        return ""


def _add_bar_chart(
    wb: Workbook,
    dim: DimensaoGrafico,
    num_rows: int,
) -> None:
    if num_rows <= 0:
        return

    ws_dash = wb[DASHBOARD_SHEET]
    ws_calc = wb[CALC_SHEET]
    _remove_charts_by_title(ws_dash, (dim.titulo,))

    end_row = 2 + num_rows
    chart = BarChart()
    chart.type = "bar" if dim.horizontal else "col"
    chart.style = 10
    chart.title = dim.titulo
    chart.y_axis.title = "Issues"
    chart.x_axis.title = dim.titulo

    data = Reference(
        ws_calc,
        min_col=dim.calc_qtde_col,
        min_row=2,
        max_row=end_row,
    )
    cats = Reference(
        ws_calc,
        min_col=dim.calc_label_col,
        min_row=3,
        max_row=end_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12 if not dim.horizontal else 14
    chart.width = 18 if not dim.horizontal else 16

    ws_dash.add_chart(chart, dim.anchor)


def _ensure_headers(wb: Workbook) -> dict:
    ws = wb[DADOS_SHEET]
    header_row, _, _ = _resolve_sheet_layout(ws)
    header_map = _build_full_header_map(ws, header_row)
    header_map = _ensure_repositorio_column(ws, header_row, header_map)
    header_map = _ensure_dev_git_columns(ws, header_row, header_map)
    return header_map


def atualizar_graficos_novos(wb: Workbook, last_data_row: int) -> dict:
    """Atualiza blocos _Calc, listas e graficos para dimensoes novas."""
    ws_dados = wb[DADOS_SHEET]
    ws_calc = wb[CALC_SHEET]
    header_map = _ensure_headers(wb)

    stats = {}
    for dim in DIMENSOES:
        col = _resolve_col(header_map, dim.header_aliases)
        if not col:
            stats[dim.key] = {"skipped": True, "reason": "coluna ausente"}
            continue

        ranked = _count_by_column(ws_dados, col, last_data_row, top_n=dim.top_n)
        labels = [label for label, _ in ranked]
        _sync_lista(wb, dim, labels)
        rows = _sync_calc_block(ws_calc, dim, ranked, col, last_data_row)
        _add_bar_chart(wb, dim, rows)
        stats[dim.key] = {
            "col": col,
            "categorias": rows,
            "total_issues": sum(q for _, q in ranked),
        }

    return stats


def sincronizar_aba_dados(wb_target: Workbook, path_source: str) -> int:
    """Copia valores da aba Dados de outro arquivo para o workbook alvo."""
    from openpyxl import load_workbook

    wb_src = load_workbook(path_source, read_only=True, data_only=True)
    ws_src = wb_src[DADOS_SHEET]
    ws_tgt = wb_target[DADOS_SHEET]
    copied = 0
    for row in ws_src.iter_rows():
        for cell in row:
            ws_tgt.cell(row=cell.row, column=cell.column, value=cell.value)
            copied += 1
    wb_src.close()
    return copied
