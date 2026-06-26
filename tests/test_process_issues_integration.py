"""Teste de integracao do nucleo de process_issues().

Exercita o caminho critico end-to-end (carregar -> filtrar -> upsert -> escrever
-> ordenar -> salvar) usando um workbook openpyxl real em arquivo temporario.

Os hooks "pesados" sao isolados de proposito:
  - detectores Git/WSL (build_*): substituidos por funcoes que retornam None,
    para nao depender de repositorios WSL nos testes;
  - hooks de dashboard (KPI, graficos, qualidade, releases, excecoes) e o save
    via COM (Excel): desligados, pois pertencem a seus proprios modulos.

Isso da uma rede de seguranca para refatorar a funcao sem alterar comportamento.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

import pytest
from openpyxl import Workbook, load_workbook

import process_gitlab_issues_v2 as pgi

HEADERS = [
    "#",
    "Título",
    "Módulo",
    "Área Funcional",
    "Criado em",
    "Estado",
    "Tipo",
    "Status",
    "Prioridade",
    "Equipe",
    "Parceria",
    "Sprint",
    "Assignee",
    "Autor",
    "Aberto?",
    "Fechado?",
]

# Hooks opcionais desligados durante o teste (cobertos pelos seus modulos).
_DISABLED_HOOKS = [
    "atualizar_kpi_parceria",
    "atualizar_graficos_novos",
    "sync_listas_taxonomia",
    "ensure_module_cols",
    "apply_module_normalization",
    "atualizar_qualidade_dados",
    "ensure_quality_columns",
    "write_confidence",
    "atualizar_releases_dashboard",
    "coletar_excecoes_wb",
    "exportar_excecoes",
    "save_workbook_preserving_filters",
]


@pytest.fixture
def dashboard_path(tmp_path: Path) -> Path:
    """Cria um MGI_Dashboard.xlsx minimo: banner na linha 1, cabecalhos na 2."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws.cell(row=1, column=1, value="MGI Dashboard - Issues")
    for col, name in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=col, value=name)
    # Issue ja existente na planilha (id 100); demais colunas vazias.
    ws.cell(row=3, column=1, value=100)
    path = tmp_path / "MGI_Dashboard.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def isolate_heavy_hooks(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(pgi, "build_detector", lambda: None)
    monkeypatch.setattr(pgi, "build_tipo_detector", lambda: None)
    monkeypatch.setattr(pgi, "build_dev_enricher", lambda: None)
    for attr in _DISABLED_HOOKS:
        monkeypatch.setattr(pgi, attr, None)


def _issue(
    issue_id: int,
    title: str,
    created: str,
    *,
    repo: str = "contratos_v2",
    state: str = "opened",
    closed: str = "",
    labels: Optional[List[str]] = None,
) -> Dict:
    return {
        "id": issue_id,
        "gitlab_repo": repo,
        "title": title,
        "createdDate": created,
        "updatedDate": created,
        "closedDate": closed,
        "state": state,
        "labels": labels or [],
        "assignees": [],
        "author": {},
        "milestone": {},
    }


def _sheet_ids(ws) -> List[int]:
    ids: List[int] = []
    for row in range(3, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value in (None, "", "#"):
            continue
        ids.append(int(value))
    return ids


@pytest.mark.usefixtures("isolate_heavy_hooks")
def test_process_issues_upsert_filtra_ordena_e_salva(dashboard_path: Path) -> None:
    recente = (datetime.now() - timedelta(days=10)).strftime("%Y-%m-%dT10:00:00")
    fechada_antiga = (datetime.now() - timedelta(days=120)).strftime("%Y-%m-%dT10:00:00")

    issues = [
        _issue(
            100,
            "[PNCP] (PNCP) - Atualiza envio",
            "2025-03-01T10:00:00",
            labels=["tipo::Melhoria", "status::Em andamento", "priority::Alta"],
        ),
        _issue(
            200,
            "[Contratos] (Gestão Contratual) - Novo recurso",
            recente,
            labels=["tipo::Bug"],
        ),
        # Antes da data de corte -> descartada (issue nova).
        _issue(50, "[API] (API v2) - Antiga", "2023-05-01T10:00:00"),
        # Fechada ha mais de 60 dias -> filtrada antes do loop.
        _issue(
            999,
            "[Jobs] (Jobs) - Fechada antiga",
            "2024-01-10T10:00:00",
            state="closed",
            closed=fechada_antiga,
        ),
    ]

    stats = pgi.process_issues(
        cutoff_date=datetime(2024, 1, 1),
        issues=issues,
        excel_file=str(dashboard_path),
        all_modules=True,
        initial_load=False,
        full_refresh=False,
    )

    assert stats is not None
    assert stats["updated_existing"] >= 1  # id 100 ja existia
    assert stats["new_added"] >= 1  # id 200 inserida
    assert stats["before_cutoff"] >= 1  # id 50 antes do corte

    wb = load_workbook(dashboard_path)
    ws = wb["Dados"]
    ids = _sheet_ids(ws)

    assert 100 in ids
    assert 200 in ids
    assert 50 not in ids  # descartada pelo filtro de data
    assert 999 not in ids  # descartada pelo filtro de fechadas (60 dias)

    # Ordenacao por # decrescente.
    assert ids == sorted(ids, reverse=True)

    # Metadados derivados do GitLab foram escritos.
    headers = {
        str(ws.cell(row=2, column=c).value): c for c in range(1, ws.max_column + 1)
    }
    estado_col = headers["Estado"]
    estados = {
        int(ws.cell(row=r, column=1).value): ws.cell(row=r, column=estado_col).value
        for r in range(3, ws.max_row + 1)
        if ws.cell(row=r, column=1).value not in (None, "", "#")
    }
    assert estados[100] == "Aberto"
    assert estados[200] == "Aberto"


@pytest.mark.usefixtures("isolate_heavy_hooks")
def test_process_issues_sem_coluna_id_retorna_none(tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    # Cabecalho sem coluna de ID (#/ID), mas com 'Título' para achar a linha.
    ws.cell(row=2, column=1, value="Título")
    ws.cell(row=2, column=2, value="Módulo")
    path = tmp_path / "MGI_Dashboard.xlsx"
    wb.save(path)

    stats = pgi.process_issues(
        cutoff_date=datetime(2024, 1, 1),
        issues=[_issue(1, "[PNCP] (PNCP) - x", "2025-01-01T10:00:00")],
        excel_file=str(path),
        all_modules=True,
    )

    assert stats is None
