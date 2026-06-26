#!/usr/bin/env python3
"""
Colunas e KPIs de qualidade de dados na aba Dados e _Calc.
"""

from __future__ import annotations

from typing import Dict, List, Optional, Tuple

from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from calc_formulas import count_by_label_formula
from modulo_normalization import CUSTOM_BUCKET
from taxonomy import assess_row_quality

DADOS_SHEET = "Dados"
CALC_SHEET = "_Calc"
DASHBOARD_SHEET = "Dashboard Executivo"
LISTAS_SHEET = "Listas"

QUALITY_FIELDS: Tuple[Tuple[str, str], ...] = (
    ("categoria", "Categoria"),
    ("modulo ok?", "Módulo OK?"),
    ("area ok?", "Área OK?"),
    ("padrao titulo?", "Padrão Título?"),
    ("padrao completo?", "Padrão Completo?"),
    ("confianca area", "Confiança Área"),
)

CALC_QUAL_LABEL_COL = 46  # AT
CALC_QUAL_QTDE_COL = 47  # AU
CALC_CAT_LABEL_COL = 52  # AZ
CALC_CAT_QTDE_COL = 53  # BA

CATEGORY_CHART_LABELS: Tuple[str, ...] = (
    "Core Business",
    "Compliance",
    "Finance",
    "Platform",
    "Operations",
    "Não mapeado",
    "Sem categoria",
)


def _normalize_header(value: str) -> str:
    import unicodedata

    text = unicodedata.normalize("NFKD", (value or "").strip())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.casefold()


def _build_header_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=col).value
        if raw:
            mapping[_normalize_header(str(raw))] = col
    return mapping


def ensure_quality_columns(ws: Worksheet, header_row: int) -> Dict[str, int]:
    """Garante colunas de qualidade no cabecalho; retorna mapa alias->coluna."""
    header_map = _build_header_map(ws, header_row)
    next_col = ws.max_column + 1
    col_map: Dict[str, int] = {}

    for alias, title in QUALITY_FIELDS:
        if alias in header_map:
            col_map[alias] = header_map[alias]
            continue
        ws.cell(row=header_row, column=next_col).value = title
        col_map[alias] = next_col
        header_map[alias] = next_col
        next_col += 1

    return col_map


def apply_quality_formulas(
    ws: Worksheet,
    header_row: int,
    data_start_row: int,
    last_row: int,
    quality_cols: Dict[str, int],
) -> int:
    """Aplica formulas de qualidade em todas as linhas de dados."""
    if last_row < data_start_row:
        return 0

    headers = _build_header_map(ws, header_row)
    mod_col = headers.get("modulo")
    norm_col = headers.get("modulo normalizado") or mod_col
    title_col = headers.get("titulo")
    area_col = headers.get("area funcional")

    mod_l = get_column_letter(norm_col) if norm_col else "C"
    title_l = get_column_letter(title_col) if title_col else "B"
    area_l = get_column_letter(area_col) if area_col else "D"

    applied = 0
    for row in range(data_start_row, last_row + 1):
        issue_id = ws.cell(row=row, column=1).value
        if issue_id in (None, "", "#"):
            continue

        cat_c = quality_cols["categoria"]
        ws.cell(row=row, column=cat_c).value = (
            f'=IF({mod_l}{row}="","",'
            f'IF({mod_l}{row}="{CUSTOM_BUCKET}","Não mapeado",'
            f'IFERROR(VLOOKUP({mod_l}{row},Lista_Modulo_Categoria,2,FALSE),"Não mapeado")))'
        )

        mod_ok_c = quality_cols["modulo ok?"]
        ws.cell(row=row, column=mod_ok_c).value = (
            f'=IF({mod_l}{row}="","Não",'
            f'IF({mod_l}{row}="{CUSTOM_BUCKET}","Não",'
            f'IF(COUNTIF(Lista_Modulos_Canonico,{mod_l}{row})>0,"Sim","Não")))'
        )

        area_ok_c = quality_cols["area ok?"]
        ws.cell(row=row, column=area_ok_c).value = (
            f'=IF({mod_l}{row}="","N/A",'
            f'IF({area_l}{row}="","Não",'
            f'IF(COUNTIF(Lista_Areas_Padrao,{area_l}{row})>0,"Sim","Não")))'
        )

        padrao_c = quality_cols["padrao titulo?"]
        ws.cell(row=row, column=padrao_c).value = (
            f'=IF({title_l}{row}="","Não",'
            f'IF(LEFT({title_l}{row},1)="[",'
            f'IF(LEN(TRIM(MID({title_l}{row},FIND("]",{title_l}{row})+1,999)))>0,"Sim","Não"),'
            f'"Não"))'
        )

        padrao_full_c = quality_cols["padrao completo?"]
        ws.cell(row=row, column=padrao_full_c).value = (
            f'=IF({title_l}{row}="","Não",'
            f'IF(ISNUMBER(SEARCH("] (",{title_l}{row})),"Sim","Não"))'
        )

        conf_c = quality_cols.get("confianca area")
        if conf_c:
            ws.cell(row=row, column=conf_c).value = (
                f'=IF({area_l}{row}="","",'
                f'IF(ISNUMBER(SEARCH("] (",{title_l}{row})),"100%","75%"))'
            )
        applied += 1

    return applied


def write_confidence(
    ws: Worksheet,
    row: int,
    conf_col: int,
    confidence: float,
    *,
    overwrite: bool = False,
) -> None:
    """Grava confianca Git (prioridade sobre formula estimada)."""
    if confidence <= 0:
        return
    current = ws.cell(row=row, column=conf_col).value
    if not overwrite and current not in (None, ""):
        if str(current).startswith("="):
            pass  # substituir formula por valor Git quando houver dado
        elif current:
            return
    ws.cell(row=row, column=conf_col).value = f"{int(round(confidence * 100))}%"


def _sync_calc_qualidade(
    ws_calc: Worksheet,
    ws_dados: Worksheet,
    quality_cols: Dict[str, int],
    last_row: int,
) -> int:
    mod_ok_col = get_column_letter(quality_cols["modulo ok?"])
    area_ok_col = get_column_letter(quality_cols["area ok?"])
    padrao_col = get_column_letter(quality_cols["padrao titulo?"])
    padrao_full_col = get_column_letter(quality_cols["padrao completo?"])
    data_range_end = last_row

    ws_calc.cell(row=1, column=CALC_QUAL_LABEL_COL).value = "qualidade"
    ws_calc.cell(row=2, column=CALC_QUAL_LABEL_COL).value = "Métrica"
    ws_calc.cell(row=2, column=CALC_QUAL_QTDE_COL).value = "Qtde"

    metrics_formulas = [
        (
            "Módulo OK",
            f'=COUNTIF(Dados!${mod_ok_col}$3:${mod_ok_col}${data_range_end},"Sim")',
        ),
        (
            "Área OK",
            f'=COUNTIF(Dados!${area_ok_col}$3:${area_ok_col}${data_range_end},"Sim")',
        ),
        (
            "Padrão Título",
            f'=COUNTIF(Dados!${padrao_col}$3:${padrao_col}${data_range_end},"Sim")',
        ),
        (
            "Padrão Completo",
            f'=COUNTIF(Dados!${padrao_full_col}$3:${padrao_full_col}${data_range_end},"Sim")',
        ),
        (
            "Total Conforme",
            f'=COUNTIFS(Dados!${mod_ok_col}$3:${mod_ok_col}${data_range_end},"Sim",'
            f'Dados!${area_ok_col}$3:${area_ok_col}${data_range_end},"Sim",'
            f'Dados!${padrao_col}$3:${padrao_col}${data_range_end},"Sim")',
        ),
        (
            "Não conformes",
            f'=MAX(0,COUNTA(Dados!$A$3:$A${data_range_end})-AU7)',
        ),
    ]

    for offset, (label, formula) in enumerate(metrics_formulas):
        row = 3 + offset
        ws_calc.cell(row=row, column=CALC_QUAL_LABEL_COL).value = label
        ws_calc.cell(row=row, column=CALC_QUAL_QTDE_COL).value = formula

    return len(metrics_formulas)


def _add_qualidade_chart(wb: Workbook, num_rows: int) -> None:
    if num_rows <= 0:
        return

    ws_dash = wb[DASHBOARD_SHEET]
    ws_calc = wb[CALC_SHEET]

    ws_dash._charts = [
        ch
        for ch in ws_dash._charts
        if not _chart_title(ch) == "Qualidade dos Dados"
    ]

    end_row = 2 + num_rows
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Qualidade dos Dados"
    chart.y_axis.title = "Issues"
    chart.x_axis.title = "Métrica"

    data = Reference(
        ws_calc,
        min_col=CALC_QUAL_QTDE_COL,
        min_row=2,
        max_row=end_row,
    )
    cats = Reference(
        ws_calc,
        min_col=CALC_QUAL_LABEL_COL,
        min_row=3,
        max_row=end_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 16

    ws_dash.cell(row=100, column=2).value = "Qualidade dos dados (conformidade)"
    ws_dash.add_chart(chart, "B101")


def _sem_categoria_formula(
    label_ref: str,
    cat_col: int,
    last_row: int,
) -> str:
    cat_letter = get_column_letter(cat_col)
    cat_range = f"Dados!${cat_letter}$3:${cat_letter}${last_row}"
    ano_range = f"Dados!$W$3:$W${last_row}"
    return (
        f'=IF({label_ref}<>"Sem categoria","",'
        f'COUNTIFS({cat_range},"",{ano_range},">="&2024))'
    )


def _sync_calc_categoria(
    ws_calc: Worksheet,
    categoria_col: int,
    last_row: int,
) -> int:
    label_letter = get_column_letter(CALC_CAT_LABEL_COL)
    ws_calc.cell(row=1, column=CALC_CAT_LABEL_COL).value = "categoria_funcional"
    ws_calc.cell(row=2, column=CALC_CAT_LABEL_COL).value = "Categoria"
    ws_calc.cell(row=2, column=CALC_CAT_QTDE_COL).value = "Qtde"

    for offset, label in enumerate(CATEGORY_CHART_LABELS):
        row = 3 + offset
        label_ref = f"{label_letter}{row}"
        ws_calc.cell(row=row, column=CALC_CAT_LABEL_COL).value = label
        if label == "Sem categoria":
            ws_calc.cell(row=row, column=CALC_CAT_QTDE_COL).value = _sem_categoria_formula(
                label_ref, categoria_col, last_row
            )
        else:
            ws_calc.cell(row=row, column=CALC_CAT_QTDE_COL).value = count_by_label_formula(
                label_ref,
                categoria_col,
                last_row,
                empty_label="Sem categoria",
            )

    return len(CATEGORY_CHART_LABELS)


def _add_categoria_chart(wb: Workbook, num_rows: int) -> None:
    if num_rows <= 0:
        return

    ws_dash = wb[DASHBOARD_SHEET]
    ws_calc = wb[CALC_SHEET]

    ws_dash._charts = [
        ch for ch in ws_dash._charts if _chart_title(ch) != "Categoria Funcional"
    ]

    end_row = 2 + num_rows
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Categoria Funcional"
    chart.y_axis.title = "Issues"
    chart.x_axis.title = "Categoria"

    data = Reference(
        ws_calc,
        min_col=CALC_CAT_QTDE_COL,
        min_row=2,
        max_row=end_row,
    )
    cats = Reference(
        ws_calc,
        min_col=CALC_CAT_LABEL_COL,
        min_row=3,
        max_row=end_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 18

    ws_dash.cell(row=118, column=2).value = "Volume por categoria funcional"
    ws_dash.add_chart(chart, "B119")


def _chart_title(chart) -> str:
    try:
        return chart.title.tx.rich.p[0].r[0].t
    except (AttributeError, IndexError, TypeError):
        return ""


def atualizar_qualidade_dados(
    wb: Workbook,
    header_row: int,
    data_start_row: int,
    last_row: int,
) -> dict:
    """Configura colunas, formulas e grafico de qualidade."""
    ws_dados = wb[DADOS_SHEET]
    quality_cols = ensure_quality_columns(ws_dados, header_row)
    applied = apply_quality_formulas(
        ws_dados, header_row, data_start_row, last_row, quality_cols
    )
    calc_rows = _sync_calc_qualidade(wb[CALC_SHEET], ws_dados, quality_cols, last_row)
    _add_qualidade_chart(wb, calc_rows)

    cat_col = quality_cols["categoria"]
    cat_rows = _sync_calc_categoria(wb[CALC_SHEET], cat_col, last_row)
    _add_categoria_chart(wb, cat_rows)

    return {
        "formulas_aplicadas": applied,
        "calc_metricas": calc_rows,
        "categoria_categorias": cat_rows,
        "colunas": quality_cols,
    }


def compute_quality_stats(
    ws: Worksheet,
    data_start_row: int,
    last_row: int,
    header_row: int,
) -> Dict[str, int]:
    """Conta conformidade via Python (para relatorio de excecoes)."""
    headers = _build_header_map(ws, header_row)
    mod_col = headers.get("modulo")
    area_col = headers.get("area funcional")
    title_col = headers.get("titulo")

    stats = {"modulo_ok": 0, "area_ok": 0, "padrao_ok": 0, "total_conforme": 0, "total": 0}
    for row in range(data_start_row, last_row + 1):
        iid = ws.cell(row=row, column=1).value
        if iid in (None, "", "#"):
            continue
        stats["total"] += 1
        title = str(ws.cell(row=row, column=title_col).value or "") if title_col else ""
        module = str(ws.cell(row=row, column=mod_col).value or "") if mod_col else ""
        area = str(ws.cell(row=row, column=area_col).value or "") if area_col else ""
        q = assess_row_quality(title, module, area)
        if q["modulo_ok"] == "Sim":
            stats["modulo_ok"] += 1
        if q["area_ok"] == "Sim":
            stats["area_ok"] += 1
        if q["padrao_titulo"] == "Sim":
            stats["padrao_ok"] += 1
        if all(
            q[k] == "Sim"
            for k in ("modulo_ok", "area_ok", "padrao_titulo")
        ):
            stats["total_conforme"] += 1
    return stats
