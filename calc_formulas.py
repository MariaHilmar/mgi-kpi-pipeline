#!/usr/bin/env python3
"""Formulas dinamicas COUNTIFS para a aba _Calc do dashboard."""

from __future__ import annotations

from openpyxl.utils import get_column_letter

DADOS_SHEET = "Dados"
DATA_START_ROW = 3
ANO_CRIACAO_COL = 23
EMPTY_LABEL = "Não informado"
OUTROS_LABEL = "Outros"
YEAR_FILTER = 2024


def _dados_range(col: int, last_row: int) -> str:
    letter = get_column_letter(col)
    return f"{DADOS_SHEET}!${letter}${DATA_START_ROW}:${letter}${last_row}"


def _ano_range(last_row: int, ano_col: int = ANO_CRIACAO_COL) -> str:
    letter = get_column_letter(ano_col)
    return f"{DADOS_SHEET}!${letter}${DATA_START_ROW}:${letter}${last_row}"


def count_by_label_formula(
    label_ref: str,
    data_col: int,
    last_row: int,
    *,
    ano_col: int = ANO_CRIACAO_COL,
    empty_label: str = EMPTY_LABEL,
) -> str:
    """COUNTIFS por rotulo na coluna de dados, filtrando Ano Criacao >= 2024."""
    data_range = _dados_range(data_col, last_row)
    ano_range = _ano_range(last_row, ano_col)
    year_criterion = f'">="&{YEAR_FILTER}'
    return (
        f'=IF({label_ref}="{empty_label}",'
        f'COUNTIFS({data_range},"",{ano_range},{year_criterion}),'
        f'COUNTIFS({data_range},{label_ref},{ano_range},{year_criterion}))'
    )


def outros_count_formula(
    qty_col_letter: str,
    row: int,
    last_row: int,
    *,
    ano_col: int = ANO_CRIACAO_COL,
) -> str:
    """Saldo de issues (ano >= 2024) menos categorias explicitas acima."""
    if row <= DATA_START_ROW:
        return "=0"
    ano_range = _ano_range(last_row, ano_col)
    year_criterion = f'">="&{YEAR_FILTER}'
    prev_sum = f"${qty_col_letter}${DATA_START_ROW}:${qty_col_letter}{row - 1}"
    return f'=MAX(0,COUNTIFS({ano_range},{year_criterion})-SUM({prev_sum}))'
