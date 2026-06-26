#!/usr/bin/env python3
"""
Aplica taxonomia, colunas de qualidade, releases e relatorio de excecoes.

Uso (sem reprocessar issues):
  python aplicar_expansao_taxonomia.py
  python aplicar_expansao_taxonomia.py --excel D:\\MGI-Relatórios\\MGI_Dashboard.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

try:
    import config
except ImportError:
    config = None

from modulo_normalization import (
    apply_module_formulas,
    apply_module_normalization,
    ensure_module_columns,
)
from atualizar_releases_dashboard import atualizar_releases_dashboard
from process_gitlab_issues_v2 import _repair_calc_formulas, _resolve_sheet_layout
from qualidade_dados import atualizar_qualidade_dados
from relatorio_excecoes import coletar_excecoes_wb, exportar


def main() -> int:
    default_excel = config.EXCEL_OUTPUT if config else Path(r"D:\MGI-Relatórios\MGI_Dashboard.xlsx")
    default_logs = config.LOGS_DIR if config else Path(r"D:\MGI-Relatórios\logs")

    parser = argparse.ArgumentParser(description="Aplica expansao taxonomia + qualidade")
    parser.add_argument("--excel", type=Path, default=default_excel)
    parser.add_argument("--output-dir", type=Path, default=default_logs)
    parser.add_argument(
        "--resync-all",
        action="store_true",
        help="Reprocessa Módulo/Original/Normalizado em linhas ja preenchidas",
    )
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"ERRO - Excel nao encontrado: {args.excel}")
        return 1

    from openpyxl import load_workbook

    print(f"OK - Carregando {args.excel}")
    wb = load_workbook(args.excel)
    ws = wb["Dados"]
    header_row, data_start, _ = _resolve_sheet_layout(ws)
    last_row = ws.max_row
    while last_row > data_start - 1 and ws.cell(row=last_row, column=1).value in (None, ""):
        last_row -= 1

    listas = sync_listas_taxonomia(wb)
    print(
        f"OK - Listas: {listas['modulos_canonicos']} canonicos, "
        f"{listas['de_para']} de-para, {listas['areas_padrao']} areas"
    )

    module_cols = ensure_module_columns(ws, header_row)
    mod_stats = apply_module_normalization(
        ws,
        header_row,
        data_start,
        last_row,
        module_cols,
        sync_modulo_column=True,
        preserve_filled_module=not args.resync_all,
    )
    preserved = mod_stats.get("preservados", 0)
    print(
        f"OK - Modulos: {mod_stats['canonicos']} normalizados, "
        f"{mod_stats['custom']} custom, {mod_stats['vazios']} vazios"
        + (f", {preserved} linhas preservadas" if preserved else "")
    )

    qualidade = atualizar_qualidade_dados(wb, header_row, data_start, last_row)
    print(f"OK - Qualidade: {qualidade['formulas_aplicadas']} formulas na aba Dados")

    releases = atualizar_releases_dashboard(wb)
    if releases["fonte_ok"]:
        print(f"OK - Releases: {releases['releases_total']} tags Git")
    else:
        print("AVISO - gitlab_git_data.json nao encontrado")

    _repair_calc_formulas(wb)

    excecoes = coletar_excecoes_wb(wb)
    paths = exportar(excecoes, args.output_dir)
    print(f"OK - {len(excecoes)} excecoes -> {paths['csv']}")

    try:
        wb.save(args.excel)
        print(f"OK - Salvo: {args.excel}")
    except PermissionError:
        alt = args.excel.with_name(f"{args.excel.stem}_taxonomia{args.excel.suffix}")
        wb.save(alt)
        print(f"AVISO - Arquivo em uso. Salvo em: {alt}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
