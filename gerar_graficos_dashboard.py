#!/usr/bin/env python3
"""
Analisa dados atualizados e gera novos graficos no MGI_Dashboard.xlsx.

Uso:
  python gerar_graficos_dashboard.py
  python gerar_graficos_dashboard.py --excel d:\\MGI-Relatórios\\MGI_Dashboard.xlsx
  python gerar_graficos_dashboard.py --fonte MGI_Dashboard_atualizado_areas.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from atualizar_dashboard_kpis import atualizar_kpi_parceria
from atualizar_graficos_dashboard import atualizar_graficos_novos, sincronizar_aba_dados
from process_gitlab_issues_v2 import _repair_calc_formulas, _resolve_sheet_layout

DEFAULT_EXCEL = Path(r"d:\MGI-Relatórios\MGI_Dashboard.xlsx")
FONTE_DADOS = Path(r"d:\MGI-Relatórios\MGI_Dashboard_atualizado_areas.xlsx")


def main() -> int:
    parser = argparse.ArgumentParser(description="Gera graficos no Dashboard Executivo")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    parser.add_argument(
        "--fonte",
        type=Path,
        default=FONTE_DADOS,
        help="Excel com Dados mais completos (Area, Desenvolvedor, etc.)",
    )
    parser.add_argument("--sem-sync", action="store_true", help="Nao copia aba Dados da fonte")
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"ERRO - Excel nao encontrado: {args.excel}")
        return 1

    from openpyxl import load_workbook

    print(f"OK - Carregando {args.excel}")
    wb = load_workbook(args.excel)

    if not args.sem_sync and args.fonte.exists() and args.fonte.resolve() != args.excel.resolve():
        print(f"OK - Sincronizando aba Dados de {args.fonte.name}")
        n = sincronizar_aba_dados(wb, str(args.fonte))
        print(f"OK - {n} celulas copiadas")

    ws = wb["Dados"]
    _, _, _ = _resolve_sheet_layout(ws)
    last_row = ws.max_row
    while last_row > 2 and ws.cell(row=last_row, column=1).value in (None, ""):
        last_row -= 1

    print(f"OK - {last_row - 2} linhas de dados")

    _repair_calc_formulas(wb)

    parceria = atualizar_kpi_parceria(wb, last_row)
    print(
        f"OK - Grafico Parceria: {parceria['parcerias']} categorias, "
        f"{parceria['formulas_updated']} formulas KPI"
    )

    stats = atualizar_graficos_novos(wb, last_row)
    for key, info in stats.items():
        if info.get("skipped"):
            print(f"AVISO - {key}: {info.get('reason')}")
        else:
            print(f"OK - Grafico {key}: {info['categorias']} categorias (col {info['col']})")

    try:
        wb.save(args.excel)
        print(f"OK - Salvo: {args.excel}")
    except PermissionError:
        alt = args.excel.with_name(f"{args.excel.stem}_graficos{args.excel.suffix}")
        wb.save(alt)
        print(f"AVISO - Arquivo em uso. Salvo em: {alt}")

    print("OK - Abra o Dashboard Executivo e pressione F9 se os KPIs nao atualizarem")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
