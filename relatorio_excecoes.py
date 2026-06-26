#!/usr/bin/env python3
"""
Exporta relatorio de excecoes de qualidade (issues nao conformes).

Uso:
  python relatorio_excecoes.py
  python relatorio_excecoes.py --excel D:\\MGI-Relatórios\\MGI_Dashboard.xlsx
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List

sys.path.insert(0, str(Path(__file__).parent))

try:
    import config
except ImportError:
    config = None

from taxonomy import assess_row_quality


def _header_map(ws, header_row: int) -> Dict[str, int]:
    import unicodedata

    def norm(value: str) -> str:
        text = unicodedata.normalize("NFKD", (value or "").strip())
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        return text.casefold()

    mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=col).value
        if raw:
            mapping[norm(str(raw))] = col
    return mapping


def coletar_excecoes(excel_path: Path) -> List[Dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, data_only=True)
    try:
        return coletar_excecoes_wb(wb)
    finally:
        wb.close()


def coletar_excecoes_wb(wb) -> List[Dict[str, str]]:
    # Import tardio para evitar dependencia circular com process_gitlab_issues_v2
    from process_gitlab_issues_v2 import _resolve_sheet_layout

    ws = wb["Dados"]
    header_row, data_start, _ = _resolve_sheet_layout(ws)
    headers = _header_map(ws, header_row)

    id_col = headers.get("#") or headers.get("id") or 1
    title_col = headers.get("titulo")
    mod_col = headers.get("modulo normalizado") or headers.get("modulo")
    area_col = headers.get("area funcional")
    repo_col = headers.get("repositorio")

    excecoes: List[Dict[str, str]] = []
    for row in range(data_start, ws.max_row + 1):
        iid = ws.cell(row=row, column=id_col).value
        if iid in (None, "", "#"):
            continue

        title = str(ws.cell(row=row, column=title_col).value or "") if title_col else ""
        module = str(ws.cell(row=row, column=mod_col).value or "") if mod_col else ""
        area = str(ws.cell(row=row, column=area_col).value or "") if area_col else ""
        repo = str(ws.cell(row=row, column=repo_col).value or "") if repo_col else ""
        orig_col = headers.get("modulo original")
        norm_col = headers.get("modulo normalizado")
        module_for_q = (
            str(ws.cell(row=row, column=norm_col).value or "")
            if norm_col
            else module
        )

        q = assess_row_quality(title, module_for_q, area)
        problemas = []
        if q["modulo_ok"] == "Não":
            problemas.append("modulo")
        if q["area_ok"] == "Não":
            problemas.append("area")
        if q["padrao_titulo"] == "Não":
            problemas.append("titulo")
        if not problemas:
            continue

        excecoes.append(
            {
                "repositorio": repo,
                "issue": str(iid),
                "titulo": title[:200],
                "modulo": module,
                "modulo_original": str(ws.cell(row=row, column=orig_col).value or "") if orig_col else "",
                "area_funcional": area,
                "modulo_ok": q["modulo_ok"],
                "area_ok": q["area_ok"],
                "padrao_titulo": q["padrao_titulo"],
                "categoria": q["categoria"],
                "problemas": ", ".join(problemas),
            }
        )

    return excecoes


def exportar(excecoes: List[Dict[str, str]], output_dir: Path) -> Dict[str, str]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = output_dir / f"excecoes_qualidade_{stamp}.csv"
    json_path = output_dir / f"excecoes_qualidade_{stamp}.json"

    fieldnames = [
        "repositorio",
        "issue",
        "titulo",
        "modulo",
        "modulo_original",
        "area_funcional",
        "modulo_ok",
        "area_ok",
        "padrao_titulo",
        "categoria",
        "problemas",
    ]

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(excecoes)

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "gerado_em": datetime.now().isoformat(),
                "total_excecoes": len(excecoes),
                "items": excecoes,
            },
            f,
            indent=2,
            ensure_ascii=False,
        )

    return {"csv": str(csv_path), "json": str(json_path)}


def main() -> int:
    default_excel = config.EXCEL_OUTPUT if config else Path(r"D:\MGI-Relatórios\MGI_Dashboard.xlsx")
    default_logs = config.LOGS_DIR if config else Path(r"D:\MGI-Relatórios\logs")

    parser = argparse.ArgumentParser(description="Relatorio de excecoes de qualidade")
    parser.add_argument("--excel", type=Path, default=default_excel)
    parser.add_argument("--output-dir", type=Path, default=default_logs)
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"ERRO - Excel nao encontrado: {args.excel}")
        return 1

    excecoes = coletar_excecoes(args.excel)
    paths = exportar(excecoes, args.output_dir)

    print(f"OK - {len(excecoes)} excecoes encontradas")
    print(f"OK - CSV: {paths['csv']}")
    print(f"OK - JSON: {paths['json']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
