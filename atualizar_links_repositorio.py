#!/usr/bin/env python3
"""
Aplica hyperlinks GitLab na coluna #, corrige Repositório e remove duplicatas.

Uso:
  python atualizar_links_repositorio.py
  python atualizar_links_repositorio.py --excel d:\\MGI-Relatórios\\MGI_Dashboard.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from issue_keys import make_issue_key
from process_gitlab_issues_v2 import (
    _build_full_header_map,
    _dedupe_dados_rows,
    _ensure_repositorio_column,
    _repair_calc_formulas,
    _resolve_sheet_layout,
    _sort_issue_rows_desc,
    _sync_issue_links_and_repos,
)

DEFAULT_EXCEL = Path(r"d:\MGI-Relatórios\MGI_Dashboard.xlsx")
ISSUES_JSON = Path(__file__).parent / "gitlab_issues_raw.json"


def main() -> int:
    parser = argparse.ArgumentParser(description="Links GitLab e dedupe na aba Dados")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL)
    args = parser.parse_args()

    if not args.excel.exists():
        print(f"ERRO - Excel nao encontrado: {args.excel}")
        return 1
    if not ISSUES_JSON.exists():
        print(f"ERRO - JSON nao encontrado: {ISSUES_JSON}")
        return 1

    with open(ISSUES_JSON, encoding="utf-8") as f:
        issues_by_id = {
            make_issue_key(i): i
            for i in json.load(f)
            if str(i.get("id", "")).strip()
        }

    from openpyxl import load_workbook

    print(f"OK - Carregando {args.excel}")
    wb = load_workbook(args.excel)
    ws = wb["Dados"]
    header_row, data_start_row, columns = _resolve_sheet_layout(ws)
    header_map = _build_full_header_map(ws, header_row)
    header_map = _ensure_repositorio_column(ws, header_row, header_map)
    repo_col = header_map.get("repositorio")
    id_col = columns["id"]

    before = ws.max_row - data_start_row + 1
    removed = _dedupe_dados_rows(
        ws,
        data_start_row,
        id_col,
        repo_col,
        columns.get("title"),
        issues_by_id,
    )
    after_dedupe = ws.max_row - data_start_row + 1

    sorted_rows = _sort_issue_rows_desc(ws, data_start_row, id_col, repo_col)
    links = _sync_issue_links_and_repos(
        ws, data_start_row, columns, repo_col, issues_by_id
    )

    _repair_calc_formulas(wb)

    try:
        wb.save(args.excel)
    except PermissionError:
        alt = args.excel.with_name(f"{args.excel.stem}_links{args.excel.suffix}")
        wb.save(alt)
        print(f"AVISO - Arquivo em uso. Salvo em {alt}")
        return 0

    print(f"OK - Linhas antes: {before}, apos dedupe: {after_dedupe}, removidas: {removed}")
    print(f"OK - {links} hyperlinks na coluna #")
    print(f"OK - {sorted_rows} linhas ordenadas")
    print(f"OK - Salvo: {args.excel}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
