#!/usr/bin/env python3
"""Sincroniza releases Git na aba _Calc e grafico no Dashboard."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl.chart import BarChart, Reference
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    import config as _config
except ImportError:
    _config = None

DASHBOARD_SHEET = "Dashboard Executivo"
CALC_SHEET = "_Calc"
LISTAS_SHEET = "Listas"

CALC_RELEASE_LABEL_COL = 48  # AV
CALC_RELEASE_QTDE_COL = 49  # AW
LISTAS_RELEASE_COL = 21  # U
TOP_RELEASES = 12


def _git_data_path() -> Path:
    if _config:
        return Path(_config.GIT_DATA_JSON)
    return Path(__file__).parent.parent / "gitlab_git_data.json"


def _parse_semver(version: str) -> Tuple[int, ...]:
    nums = re.findall(r"\d+", version or "")
    if not nums:
        return (0,)
    return tuple(int(n) for n in nums[:4])


def _load_releases(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    releases: List[Dict[str, str]] = []

    # Formato consolidado multi-repo
    if isinstance(data.get("repositorios"), list):
        for repo_block in data["repositorios"]:
            repo_name = repo_block.get("repositorio", "")
            for rel in repo_block.get("releases") or []:
                releases.append(
                    {
                        "versao": rel.get("versao", ""),
                        "data": rel.get("data", ""),
                        "repositorio": repo_name,
                        "rotulo": f"{repo_name}: {rel.get('versao', '')}",
                    }
                )
        return releases

    # Formato legado repo unico
    repo_name = data.get("repositorio", "contratos_v2")
    for rel in data.get("releases") or []:
        releases.append(
            {
                "versao": rel.get("versao", ""),
                "data": rel.get("data", ""),
                "repositorio": repo_name,
                "rotulo": f"{repo_name}: {rel.get('versao', '')}",
            }
        )
    return releases


def _rank_releases(releases: List[Dict[str, str]]) -> List[Dict[str, str]]:
    return sorted(
        releases,
        key=lambda r: (_parse_semver(r.get("versao", "")), r.get("data", "")),
        reverse=True,
    )[:TOP_RELEASES]


def _sync_lista_releases(wb: Workbook, releases: List[Dict[str, str]]) -> None:
    ws = wb[LISTAS_SHEET]
    ws.cell(row=1, column=LISTAS_RELEASE_COL).value = "Release Git"
    ws.cell(row=2, column=LISTAS_RELEASE_COL).value = "Todos"
    for idx, rel in enumerate(releases, start=3):
        ws.cell(row=idx, column=LISTAS_RELEASE_COL).value = rel.get("versao", "")


def _sync_calc_releases(ws_calc: Worksheet, releases: List[Dict[str, str]]) -> int:
    ws_calc.cell(row=1, column=CALC_RELEASE_LABEL_COL).value = "releases"
    ws_calc.cell(row=2, column=CALC_RELEASE_LABEL_COL).value = "Release"
    ws_calc.cell(row=2, column=CALC_RELEASE_QTDE_COL).value = "Data"

    for offset, rel in enumerate(releases):
        row = 3 + offset
        ws_calc.cell(row=row, column=CALC_RELEASE_LABEL_COL).value = rel.get("rotulo", "")
        ws_calc.cell(row=row, column=CALC_RELEASE_QTDE_COL).value = rel.get("data", "")

    for row in range(3 + len(releases), ws_calc.max_row + 1):
        if ws_calc.cell(row=row, column=CALC_RELEASE_LABEL_COL).value:
            ws_calc.cell(row=row, column=CALC_RELEASE_LABEL_COL).value = None
            ws_calc.cell(row=row, column=CALC_RELEASE_QTDE_COL).value = None

    return len(releases)


def _chart_title(chart) -> str:
    try:
        return chart.title.tx.rich.p[0].r[0].t
    except (AttributeError, IndexError, TypeError):
        return ""


def _add_releases_chart(wb: Workbook, releases: List[Dict[str, str]]) -> None:
    if not releases:
        return

    ws_dash = wb[DASHBOARD_SHEET]
    ws_calc = wb[CALC_SHEET]

    ws_dash._charts = [
        ch for ch in ws_dash._charts if _chart_title(ch) != "Releases Git"
    ]

    # Contagem por repositorio (formula estatica escrita como valor — atualizada a cada pipeline)
    by_repo: Dict[str, int] = {}
    for rel in releases:
        repo = rel.get("repositorio", "?")
        by_repo[repo] = by_repo.get(repo, 0) + 1

    ws_calc.cell(row=1, column=CALC_RELEASE_LABEL_COL + 2).value = "releases_repo"
    ws_calc.cell(row=2, column=CALC_RELEASE_LABEL_COL + 2).value = "Repositório"
    ws_calc.cell(row=2, column=CALC_RELEASE_LABEL_COL + 3).value = "Tags"

    for offset, (repo, count) in enumerate(sorted(by_repo.items())):
        row = 3 + offset
        ws_calc.cell(row=row, column=CALC_RELEASE_LABEL_COL + 2).value = repo
        ws_calc.cell(row=row, column=CALC_RELEASE_LABEL_COL + 3).value = count

    end_row = 2 + len(by_repo)
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Releases Git"
    chart.y_axis.title = "Tags (top)"
    chart.x_axis.title = "Repositório"

    data = Reference(
        ws_calc,
        min_col=CALC_RELEASE_LABEL_COL + 3,
        min_row=2,
        max_row=end_row,
    )
    cats = Reference(
        ws_calc,
        min_col=CALC_RELEASE_LABEL_COL + 2,
        min_row=3,
        max_row=end_row,
    )
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 14

    ws_dash.cell(row=100, column=10).value = "Releases Git (tags recentes)"
    ws_dash.add_chart(chart, "J101")


def atualizar_releases_dashboard(
    wb: Workbook,
    git_json: Optional[Path] = None,
) -> dict:
    path = git_json or _git_data_path()
    all_releases = _load_releases(path)
    ranked = _rank_releases(all_releases)

    _sync_lista_releases(wb, ranked)
    rows = _sync_calc_releases(wb[CALC_SHEET], ranked)
    _add_releases_chart(wb, all_releases)

    return {
        "git_json": str(path),
        "releases_total": len(all_releases),
        "releases_exibidas": rows,
        "fonte_ok": path.exists(),
    }
