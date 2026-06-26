#!/usr/bin/env python3
"""
Preenche a coluna Desenvolvedor na aba Dados.

Prioridade:
  1. Autor com mais commits na branch {iid}-* (Git/WSL)
  2. Autor em commits que referenciam a issue (#iid, {iid}-)
  3. Primeiro assignee da issue no GitLab

Uso:
  python preencher_desenvolvedor.py
  python preencher_desenvolvedor.py --force
  python preencher_desenvolvedor.py --excel MGI_Dashboard_atualizado_areas.xlsx
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

try:
    import config as _config
except ImportError:
    _config = None

from enriquecer_dev_git import build_dev_enricher
from issue_keys import make_issue_key
from process_gitlab_issues_v2 import (
    _backfill_desenvolvedor_for_sheet,
    _build_full_header_map,
    _ensure_dev_git_columns,
    _ensure_repositorio_column,
    _repair_calc_formulas,
    _resolve_sheet_layout,
    _row_issue_key,
)

DEFAULT_XLSX = Path(r"D:\MGI-Relatórios\MGI_Dashboard_atualizado_areas.xlsx")
FALLBACK_XLSX = Path(r"D:\MGI-Relatórios\MGI_Dashboard_atualizado.xlsx")
ISSUES_JSON = (
    Path(_config.ISSUES_JSON)
    if _config
    else Path(__file__).parent / "gitlab_issues_raw.json"
)


def _pick_excel(path: str | None) -> Path:
    if path:
        return Path(path)
    if DEFAULT_XLSX.exists():
        return DEFAULT_XLSX
    return FALLBACK_XLSX


def main() -> int:
    parser = argparse.ArgumentParser(description="Preenche Desenvolvedor no MGI_Dashboard")
    parser.add_argument("--excel", help="Caminho do Excel")
    parser.add_argument("--force", action="store_true", help="Reprocessa linhas ja preenchidas")
    args = parser.parse_args()

    excel_path = _pick_excel(args.excel)
    if not excel_path.exists():
        print(f"ERRO - Excel nao encontrado: {excel_path}")
        return 1
    if not ISSUES_JSON.exists():
        print(f"ERRO - JSON nao encontrado: {ISSUES_JSON}")
        return 1

    with open(ISSUES_JSON, encoding="utf-8") as f:
        issues_by_id = {
            make_issue_key(issue): issue
            for issue in json.load(f)
            if str(issue.get("id", "")).strip()
        }

    from openpyxl import load_workbook

    print(f"OK - Carregando {excel_path}", flush=True)
    wb = load_workbook(excel_path)
    ws = wb["Dados"]
    header_row, data_start_row, columns = _resolve_sheet_layout(ws)
    header_map = _build_full_header_map(ws, header_row)
    header_map = _ensure_repositorio_column(ws, header_row, header_map)
    header_map = _ensure_dev_git_columns(ws, header_row, header_map)
    repo_col = header_map.get("repositorio")
    id_col = columns["id"]
    dev_col = header_map.get("desenvolvedor")

    enricher = build_dev_enricher()

    empty_before = 0
    rows = 0
    for row in range(data_start_row, ws.max_row + 1):
        if not _row_issue_key(ws, row, id_col, repo_col):
            continue
        rows += 1
        if ws.cell(row=row, column=dev_col).value in (None, ""):
            empty_before += 1

    print(f"OK - {rows} linhas, {empty_before} sem Desenvolvedor")
    if args.force:
        print("OK - Modo --force", flush=True)

    filled = _backfill_desenvolvedor_for_sheet(
        ws,
        header_map,
        data_start_row,
        id_col,
        repo_col,
        issues_by_id,
        enricher,
        force=args.force,
    )

    empty_after = 0
    for row in range(data_start_row, ws.max_row + 1):
        if not _row_issue_key(ws, row, id_col, repo_col):
            continue
        if ws.cell(row=row, column=dev_col).value in (None, ""):
            empty_after += 1

    _repair_calc_formulas(wb)

    try:
        wb.save(excel_path)
    except PermissionError:
        alt = excel_path.with_name(f"{excel_path.stem}_dev{excel_path.suffix}")
        print(f"AVISO - Arquivo em uso. Gravando em {alt}")
        wb.save(alt)
        excel_path = alt

    print(f"OK - {filled} desenvolvedores preenchidos")
    print(f"OK - Vazios: {empty_before} -> {empty_after}")
    print(f"OK - Salvo: {excel_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
