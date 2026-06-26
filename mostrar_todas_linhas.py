#!/usr/bin/env python3
"""
Script para mostrar todas as linhas ocultas no Excel
"""

from openpyxl import load_workbook

# Abrir o arquivo
excel_file = r'D:\MGI-Relatórios\MGI_Dashboard.xlsx'
wb = load_workbook(excel_file)
ws = wb['Dados']

print("Mostrando todas as linhas ocultas...")

# Mostrar todas as linhas
for row in ws.iter_rows():
    ws.row_dimensions[row[0].row].hidden = False

# Salvar
wb.save(excel_file)
print(f"✓ Arquivo salvo com todas as linhas visíveis: {excel_file}")
print(f"✓ Total de linhas na planilha: {ws.max_row}")
