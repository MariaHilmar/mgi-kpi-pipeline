#!/usr/bin/env python3
"""Compara Dados-(ajustes manuais).xlsx com MGI_Dashboard.xlsx."""
from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
import json
import unicodedata

from openpyxl import load_workbook


def norm(s: str) -> str:
    t = unicodedata.normalize("NFKD", str(s or "").strip())
    return "".join(c for c in t if not unicodedata.combining(c)).casefold()


def load_sheet(path: Path, header_row: int, data_start: int) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Dados"]
    hm: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h:
            hm[norm(str(h))] = c

    id_col = hm.get("#") or hm.get("id") or 1
    rows: list[dict] = []
    for r in range(data_start, ws.max_row + 1):
        iid = ws.cell(r, id_col).value
        if iid in (None, "", "#"):
            continue
        rows.append(
            {
                "id": str(iid).strip(),
                "titulo": str(ws.cell(r, hm["titulo"]).value or ""),
                "modulo": str(ws.cell(r, hm["modulo"]).value or "").strip(),
                "area": str(ws.cell(r, hm["area funcional"]).value or "").strip(),
            }
        )
    wb.close()
    return rows


def main() -> None:
    manual_path = Path(r"D:\MGI-Relatórios\Dados-(ajustes manuais).xlsx")
    dash_path = Path(r"D:\MGI-Relatórios\MGI_Dashboard.xlsx")
    logs = Path(r"D:\MGI-Relatórios\logs")
    logs.mkdir(exist_ok=True)

    manual = load_sheet(manual_path, header_row=1, data_start=2)
    current = load_sheet(dash_path, header_row=2, data_start=3)

    cur_by_id = {r["id"]: r for r in current}
    man_by_id = {r["id"]: r for r in manual}
    common = set(cur_by_id) & set(man_by_id)

    mod_changes: list[dict] = []
    area_changes: list[dict] = []
    both_changes: list[dict] = []

    for iid in common:
        m, c = man_by_id[iid], cur_by_id[iid]
        mod_diff = m["modulo"] != c["modulo"]
        area_diff = m["area"] != c["area"]
        if not mod_diff and not area_diff:
            continue
        rec = {
            "id": iid,
            "titulo": m["titulo"][:200],
            "manual_mod": m["modulo"],
            "curr_mod": c["modulo"],
            "manual_area": m["area"],
            "curr_area": c["area"],
        }
        if mod_diff and area_diff:
            both_changes.append(rec)
        elif mod_diff:
            mod_changes.append(rec)
        else:
            area_changes.append(rec)

    # De-para agregado: valor atual -> valor manual
    mod_depara: Counter = Counter()
    area_depara: Counter = Counter()
    mod_area_pairs_manual: Counter = Counter()
    tag_to_manual_mod: Counter = Counter()

    import re

    tag_re = re.compile(r"^\[([^\]]+)\]")

    for iid in common:
        m = man_by_id[iid]
        c = cur_by_id[iid]
        if m["modulo"] != c["modulo"]:
            mod_depara[(c["modulo"], m["modulo"])] += 1
        if m["area"] != c["area"]:
            area_depara[(c["area"], m["area"])] += 1
        if m["modulo"] or m["area"]:
            mod_area_pairs_manual[(m["modulo"], m["area"])] += 1
        match = tag_re.match(m["titulo"])
        if match and m["modulo"]:
            tag_to_manual_mod[(match.group(1).strip(), m["modulo"])] += 1

    report = {
        "manual_rows": len(manual),
        "current_rows": len(current),
        "common_ids": len(common),
        "modulo_changes": len(mod_changes),
        "area_changes": len(area_changes),
        "both_changes": len(both_changes),
        "manual_modules": Counter(r["modulo"] for r in manual if r["modulo"]),
        "manual_areas": Counter(r["area"] for r in manual if r["area"]),
        "current_modules": Counter(r["modulo"] for r in current if r["modulo"]),
        "current_areas": Counter(r["area"] for r in current if r["area"]),
        "mod_depara_top": mod_depara.most_common(40),
        "area_depara_top": area_depara.most_common(40),
        "tag_to_manual_mod_top": tag_to_manual_mod.most_common(50),
        "mod_area_pairs_manual_top": mod_area_pairs_manual.most_common(40),
    }

    out_json = logs / "analise_ajustes_manuais.json"
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(
            {
                k: (dict(v) if isinstance(v, Counter) else v)
                for k, v in report.items()
            },
            f,
            ensure_ascii=False,
            indent=2,
        )

    # CSVs
    import csv

    for name, items in [
        ("mudancas_modulo", mod_changes + both_changes),
        ("mudancas_area", area_changes + both_changes),
    ]:
        p = logs / f"ajustes_manuais_{name}.csv"
        with open(p, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(
                f,
                fieldnames=[
                    "id",
                    "titulo",
                    "curr_mod",
                    "manual_mod",
                    "curr_area",
                    "manual_area",
                ],
            )
            w.writeheader()
            w.writerows(items)

    print(f"manual rows: {len(manual)}")
    print(f"current rows: {len(current)}")
    print(f"ids em comum: {len(common)}")
    print(f"modulo alterado: {len(mod_changes)} (+ {len(both_changes)} com area)")
    print(f"area alterada: {len(area_changes)} (+ {len(both_changes)} com modulo)")
    print(f"\nJSON: {out_json}")
    print("\nTop mudancas MODULO (atual -> manual):")
    for (a, b), n in mod_depara.most_common(25):
        print(f"  {n:4d}  [{a}] -> [{b}]")
    print("\nTop mudancas AREA (atual -> manual):")
    for (a, b), n in area_depara.most_common(25):
        print(f"  {n:4d}  ({a}) -> ({b})")
    print("\nTop tags [X] -> modulo manual:")
    for (tag, mod), n in tag_to_manual_mod.most_common(25):
        print(f"  {n:4d}  [{tag}] -> {mod}")


if __name__ == "__main__":
    main()
