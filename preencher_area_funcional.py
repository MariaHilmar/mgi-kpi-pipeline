#!/usr/bin/env python3
"""
Repreenche a coluna Area Funcional na aba Dados.

Prioridade de deteccao (sem --titulo-only):
  1. Area explicita no titulo: [Modulo] (Area) - ...
  2. Arquivos alterados em branches/commits Git ({iid}-*, #iid)
  3. Palavras-chave no nome da branch
  4. Palavras-chave no titulo
  5. Modulo [X] no titulo

Uso:
  python preencher_area_funcional.py
  python preencher_area_funcional.py --force
  python preencher_area_funcional.py --force --only-area "Portal Fornecedor"
  python preencher_area_funcional.py --titulo-only

Variaveis de ambiente:
  CMD:     set MGI_AREA_SKIP_GIT_GREP=1
  PowerShell: $env:MGI_AREA_SKIP_GIT_GREP='1'
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

try:
    import config as _config
except ImportError:
    _config = None

from detectar_area_funcional import build_detector
from issue_keys import make_issue_key
from process_gitlab_issues_v2 import (
    _backfill_areas_for_sheet,
    _build_full_header_map,
    _ensure_repositorio_column,
    _repair_calc_formulas,
    _resolve_sheet_layout,
    _row_issue_key,
)

DEFAULT_XLSX = (
    Path(_config.EXCEL_OUTPUT).parent / "MGI_Dashboard_atualizado.xlsx"
    if _config
    else Path(r"D:\MGI-Relatórios\MGI_Dashboard_atualizado.xlsx")
)
FALLBACK_XLSX = (
    Path(_config.EXCEL_OUTPUT)
    if _config
    else Path(r"D:\MGI-Relatórios\MGI_Dashboard.xlsx")
)
AREAS_XLSX = Path(r"D:\MGI-Relatórios\MGI_Dashboard_atualizado_areas.xlsx")
ISSUES_JSON = (
    Path(_config.ISSUES_JSON)
    if _config
    else Path(__file__).parent / "gitlab_issues_raw.json"
)


def _pick_excel(path: str | None) -> Path:
    if path:
        return Path(path)
    if AREAS_XLSX.exists():
        return AREAS_XLSX
    if DEFAULT_XLSX.exists():
        return DEFAULT_XLSX
    return FALLBACK_XLSX


def main() -> int:
    parser = argparse.ArgumentParser(description="Preenche Area Funcional no MGI_Dashboard")
    parser.add_argument("--excel", help="Caminho do Excel")
    parser.add_argument("--force", action="store_true", help="Reprocessa linhas (todas ou --only-area)")
    parser.add_argument(
        "--only-area",
        help='So corrige linhas com esta area (ex.: "Portal Fornecedor"; requer --force)',
    )
    parser.add_argument(
        "--titulo-only",
        action="store_true",
        help="Ignora Git — rapido, mas menos preciso",
    )
    args = parser.parse_args()

    if args.only_area and not args.force:
        print("ERRO - --only-area requer --force")
        return 1

    if args.titulo_only:
        os.environ["MGI_AREA_TITULO_ONLY"] = "1"
    else:
        os.environ.pop("MGI_AREA_TITULO_ONLY", None)

    excel_path = _pick_excel(args.excel)
    if not excel_path.exists():
        print(f"ERRO - Excel nao encontrado: {excel_path}")
        return 1
    if not ISSUES_JSON.exists():
        print(f"ERRO - JSON nao encontrado: {ISSUES_JSON}")
        return 1

    with open(ISSUES_JSON, encoding="utf-8") as f:
        issues = json.load(f)

    issues_by_id = {
        make_issue_key(issue): issue
        for issue in issues
        if str(issue.get("id", "")).strip()
    }

    from openpyxl import load_workbook

    mode = "titulo-only" if args.titulo_only else "git+titulo"
    print(f"OK - Carregando {excel_path}", flush=True)
    print(f"OK - Modo: {mode}", flush=True)

    wb = load_workbook(excel_path)
    ws = wb["Dados"]
    header_row, data_start_row, columns = _resolve_sheet_layout(ws)
    header_map = _build_full_header_map(ws, header_row)
    header_map = _ensure_repositorio_column(ws, header_row, header_map)
    repo_col = header_map.get("repositorio")
    id_col = columns["id"]

    detector = build_detector()
    if not detector:
        print("ERRO - Detector de area funcional indisponivel")
        return 1

    empty_before = 0
    rows = 0
    area_col = columns["area"]
    for row in range(data_start_row, ws.max_row + 1):
        if not _row_issue_key(ws, row, id_col, repo_col):
            continue
        rows += 1
        if ws.cell(row=row, column=area_col).value in (None, ""):
            empty_before += 1

    print(f"OK - {rows} linhas, {empty_before} com Area Funcional vazia")
    if args.force and args.only_area:
        print(f"OK - Corrigindo apenas area: {args.only_area}", flush=True)
    elif args.force:
        print("OK - Modo --force: reprocessando todas as linhas", flush=True)

    filled = _backfill_areas_for_sheet(
        ws,
        columns,
        data_start_row,
        id_col,
        repo_col,
        issues_by_id,
        detector,
        force=args.force,
        only_area=args.only_area,
    )

    empty_after = 0
    for row in range(data_start_row, ws.max_row + 1):
        if not _row_issue_key(ws, row, id_col, repo_col):
            continue
        if ws.cell(row=row, column=area_col).value in (None, ""):
            empty_after += 1

    _repair_calc_formulas(wb)

    try:
        wb.save(excel_path)
    except PermissionError:
        alt = excel_path.with_name(f"{excel_path.stem}_areas{excel_path.suffix}")
        print(f"AVISO - Arquivo em uso. Gravando em {alt}")
        wb.save(alt)
        excel_path = alt

    print(f"OK - {filled} areas preenchidas/corrigidas")
    print(f"OK - Vazias: {empty_before} -> {empty_after}")
    print(f"OK - Salvo: {excel_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
