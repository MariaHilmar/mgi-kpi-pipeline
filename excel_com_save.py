#!/usr/bin/env python3
"""
Grava apenas a aba Dados via Excel COM (Windows), preservando:
- aba Listas e valores dos filtros
- validacoes de dados (dropdowns)
- demais abas do dashboard (Dashboard Executivo, Alertas, etc.)
"""

from __future__ import annotations

import os
import time
from datetime import date, datetime
from typing import List, Sequence, Tuple

DATOS_SHEET = "Dados"
ISSUES_TABLE = "tbIssues"
ID_COLUMN = 1
# RGB(255, 192, 0) -> laranja Excel (BGR) — tipo inferido
INFERRED_TIPO_EXCEL_COLOR = 49407
# RGB(189, 215, 238) -> azul claro Excel (BGR) — dados Git
DEV_GIT_EXCEL_COLOR = 15652797

PROTECTED_SHEETS = frozenset({
    "Listas",
    "Dashboard Executivo",
    "Análise Temporal",
    "Detalhamento",
    "🚨 Alertas",
    "_Calc",
})


def com_available() -> bool:
    try:
        import win32com.client  # noqa: F401
        return True
    except ImportError:
        return False


def _to_com_value(value):
    """Converte tipos Python para valores aceitos pelo Excel COM."""
    if value is None:
        return None
    if isinstance(value, datetime):
        import pywintypes
        return pywintypes.Time(value.replace(tzinfo=None))
    if isinstance(value, date):
        import pywintypes
        return pywintypes.Time(datetime.combine(value, datetime.min.time()))
    return value


def _normalize_id(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def _count_ids_com(ws, data_start_row: int, last_row: int) -> int:
    count = 0
    for row in range(data_start_row, last_row + 1):
        issue_id = _normalize_id(ws.Cells(row, ID_COLUMN).Value)
        if issue_id and issue_id != "#":
            count += 1
    return count


def _count_ids_openpyxl(excel_path: str, data_start_row: int, last_row: int) -> int:
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = wb[DATOS_SHEET]
        count = 0
        for row in range(data_start_row, last_row + 1):
            issue_id = _normalize_id(ws.cell(row=row, column=ID_COLUMN).value)
            if issue_id and issue_id != "#":
                count += 1
        return count
    finally:
        wb.close()


def save_dados_sheet_via_com(
    excel_path: str,
    header_row: int,
    data_start_row: int,
    last_row: int,
    last_col: int,
    cell_values,
    expected_rows: int | None = None,
    highlight_cells: Sequence[Tuple[int, int] | Tuple[int, int, int]] | None = None,
) -> int:
    """Persiste valores da aba Dados usando Excel instalado (nao altera outras abas)."""
    path = os.path.abspath(excel_path)
    last_error = None

    for attempt in range(1, 4):
        try:
            return _save_dados_sheet_via_com_attempt(
                path,
                header_row,
                data_start_row,
                last_row,
                last_col,
                cell_values,
                expected_rows,
                highlight_cells,
            )
        except PermissionError:
            raise
        except Exception as exc:
            last_error = exc
            if attempt < 3 and getattr(exc, "args", None) and exc.args[0] == -2147418111:
                time.sleep(2)
                continue
            raise

    if last_error:
        raise last_error
    return 0


def _save_dados_sheet_via_com_attempt(
    path: str,
    header_row: int,
    data_start_row: int,
    last_row: int,
    last_col: int,
    cell_values,
    expected_rows: int | None,
    highlight_cells: Sequence[Tuple[int, int] | Tuple[int, int, int]] | None = None,
) -> int:
    import win32com.client

    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.ScreenUpdating = False

    try:
        wb = excel.Workbooks.Open(Filename=path, UpdateLinks=0, ReadOnly=False)
        if wb.ReadOnly:
            wb.Close(SaveChanges=False)
            raise PermissionError(
                f"Arquivo aberto ou somente leitura. Feche o Excel antes de gravar: {path}"
            )

        ws = wb.Worksheets(DATOS_SHEET)

        # Expande a tabela antes de gravar novas linhas
        _resize_issues_table(ws, header_row, last_row, last_col)

        row_count = last_row - data_start_row + 1
        if row_count > 0 and last_col > 0:
            matrix = [
                [
                    _to_com_value(cell_values(data_start_row + offset, col))
                    for col in range(1, last_col + 1)
                ]
                for offset in range(row_count)
            ]
            target = ws.Range(
                ws.Cells(data_start_row, 1),
                ws.Cells(last_row, last_col),
            )
            target.Value = matrix

        if highlight_cells:
            for item in highlight_cells:
                if len(item) >= 3:
                    row, col, color = item[0], item[1], item[2]
                else:
                    row, col = item[0], item[1]
                    color = INFERRED_TIPO_EXCEL_COLOR
                ws.Cells(row, col).Interior.Color = color

        used_last = ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1
        if used_last > last_row:
            clear_range = ws.Range(
                ws.Cells(last_row + 1, 1),
                ws.Cells(used_last, last_col),
            )
            clear_range.ClearContents()

        _resize_issues_table(ws, header_row, last_row, last_col)

        saved_rows = _count_ids_com(ws, data_start_row, last_row)
        if expected_rows is not None and saved_rows < expected_rows:
            wb.Close(SaveChanges=False)
            raise RuntimeError(
                f"Gravacao incompleta via COM: {saved_rows} linhas salvas, "
                f"esperado {expected_rows}"
            )

        wb.Save()
        wb.Close(SaveChanges=True)
    finally:
        try:
            excel.Quit()
        except Exception:
            pass

    disk_rows = _count_ids_openpyxl(path, data_start_row, last_row)
    if expected_rows is not None and disk_rows < expected_rows:
        raise RuntimeError(
            f"Arquivo gravado incompleto: {disk_rows} linhas no disco, "
            f"esperado {expected_rows}. Feche o Excel e tente novamente."
        )

    return disk_rows


def _resize_issues_table(ws, header_row: int, last_row: int, last_col: int) -> None:
    try:
        table = ws.ListObjects(ISSUES_TABLE)
    except Exception:
        return

    new_range = ws.Range(
        ws.Cells(header_row, 1),
        ws.Cells(last_row, last_col),
    )
    table.Resize(new_range)


def save_workbook_preserving_filters(
    excel_path: str,
    ws,
    header_row: int,
    data_start_row: int,
    last_data_row: int | None = None,
    expected_data_rows: int | None = None,
    highlight_cells: Sequence[Tuple[int, int] | Tuple[int, int, int]] | None = None,
) -> bool:
    """Tenta salvar via COM. Retorna True se conseguiu."""
    if not com_available():
        return False

    last_row = last_data_row or ws.max_row
    last_col = ws.max_column

    def getter(row: int, col: int):
        return ws.cell(row=row, column=col).value

    try:
        saved_rows = save_dados_sheet_via_com(
            excel_path,
            header_row=header_row,
            data_start_row=data_start_row,
            last_row=last_row,
            last_col=last_col,
            cell_values=getter,
            expected_rows=expected_data_rows,
            highlight_cells=highlight_cells,
        )
        print("OK - Gravacao via Excel COM (Listas e filtros preservados)")
        print(f"OK - Linhas com issue na aba Dados: {saved_rows}")
        return True
    except PermissionError as exc:
        print(f"AVISO - Falha ao gravar via Excel COM: {exc}")
        return False
    except Exception as exc:
        print(f"AVISO - Falha ao gravar via Excel COM: {exc}")
        return False
