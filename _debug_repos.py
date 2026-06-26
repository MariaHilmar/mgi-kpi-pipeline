#!/usr/bin/env python3
import json
from collections import Counter
from datetime import datetime

from process_gitlab_issues_v2 import (
    parse_date,
    extract_module,
    make_issue_key,
    _modulo_permitido,
    _normalize_id,
    _row_issue_key,
)

JSON_PATH = r"D:\MGI-Relatórios\mgi\gitlab_issues_raw.json"
XLSX_PATH = r"D:\MGI-Relatórios\MGI_Dashboard.xlsx"


def main():
    with open(JSON_PATH, encoding="utf-8") as f:
        issues = json.load(f)

    from openpyxl import load_workbook

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb["Dados"]
    repo_col = 41  # coluna Repositório conhecida
    id_col = 1

    existing_keys = set()
    repo_counts = Counter()
    for r in range(3, ws.max_row + 1):
        key = _row_issue_key(ws, r, id_col, repo_col)
        if key:
            existing_keys.add(key)
            repo_counts[key.split(":")[0]] += 1

    print("Excel existing keys:", len(existing_keys))
    print("Excel por repo (chave):", dict(repo_counts))

    cutoff = datetime(2024, 1, 1)
    would_add = Counter()
    for issue in issues:
        key = make_issue_key(issue)
        created = parse_date(issue.get("createdDate", ""))
        if created and created < cutoff:
            continue
        if key in existing_keys:
            continue
        if not _modulo_permitido(extract_module(issue.get("title", "")), all_modules=True):
            continue
        would_add[issue.get("gitlab_repo", "?")] += 1

    print("Faltam inserir (apos 2024, all_modules):", dict(would_add))
    print("Total faltando:", sum(would_add.values()))
    wb.close()


if __name__ == "__main__":
    main()
