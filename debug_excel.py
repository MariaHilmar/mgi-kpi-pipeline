#!/usr/bin/env python3
"""
Script para debugar o que está no Excel
"""

from openpyxl import load_workbook

excel_file = r'D:\MGI-Relatórios\MGI_Dashboard.xlsx'
wb = load_workbook(excel_file)
ws = wb['Dados']

print("\n" + "="*70)
print("DEBUG - EXCEL")
print("="*70)

print(f"Total de linhas com dados: {ws.max_row}")
print(f"Total de colunas: {ws.max_column}")

# Mostrar últimas 20 linhas
print("\n✓ ÚLTIMAS 20 LINHAS:")
print("-"*70)

start_row = max(2, ws.max_row - 20)
for row in range(start_row, ws.max_row + 1):
    id_val = ws.cell(row=row, column=1).value  # ID
    title = ws.cell(row=row, column=2).value   # Título
    module = ws.cell(row=row, column=3).value  # Módulo

    if title:
        print(f"Linha {row}: ID={id_val} | Módulo={module} | Título={title[:60]}")

# Contar por módulo
print("\n✓ CONTAGEM POR MÓDULO:")
print("-"*70)

modulos = {}
for row in range(2, ws.max_row + 1):
    module = ws.cell(row=row, column=3).value
    if module:
        modulos[module] = modulos.get(module, 0) + 1

for modulo, count in sorted(modulos.items(), key=lambda x: -x[1]):
    print(f"  {modulo}: {count}")

print(f"\n✓ Total de registros: {ws.max_row - 1}")
print("="*70)
