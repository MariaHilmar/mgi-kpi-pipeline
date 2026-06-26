#!/usr/bin/env python3
"""
Sincroniza a aba Dados do Excel (e releases Git) para o Supabase.

Uso:
  python sync_supabase.py
  python sync_supabase.py --excel D:\\MGI-Relatórios\\MGI_Dashboard.xlsx
"""

from __future__ import annotations

import argparse
import json
import os
import unicodedata
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    import config as _config
except ImportError:
    _config = None

DADOS_SHEET = "Dados"

# Excel header (normalizado) -> coluna Supabase
COLUMN_MAP: Dict[str, str] = {
    "#": "gitlab_iid",
    "id": "gitlab_iid",
    "titulo": "titulo",
    "modulo": "modulo",
    "modulo normalizado": "modulo_normalizado",
    "area funcional": "area_funcional",
    "tipo": "tipo",
    "estado": "estado",
    "status": "status",
    "prioridade": "prioridade",
    "equipe": "equipe",
    "parceria": "parceria",
    "sprint": "sprint",
    "assignee": "assignee",
    "autor": "autor",
    "solicitante": "solicitante",
    "alteracao escopo": "alteracao_escopo",
    "repositorio": "repositorio",
    "desenvolvedor": "desenvolvedor",
    "criado em": "criado_em",
    "data criacao": "criado_em",
    "fechado em": "fechado_em",
    "lead time": "lead_time_dias",
    "ano/mes criacao": "ano_mes_criacao",
    "ano criacao": "ano_criacao",
    "mes criacao data": "mes_criacao",
    "ano/mes fechamento": "ano_mes_fechamento",
    "mes fechamento data": "mes_fechamento",
    "aberto?": "aberto",
    "fechado?": "fechado",
    "idade (dias)": "idade_dias",
    "sla > 90 dias": "sla_mais_90_dias",
    "dev: tem branch": "dev_tem_branch",
    "dev: branch": "dev_branch",
    "dev: commits": "dev_commits",
    "dev: ultimo commit": "dev_ultimo_commit",
    "dev: autor dev": "dev_autor_dev",
    "gitlab: mrs": "gitlab_mrs",
    "dev: mergeado?": "dev_mergeado",
    "categoria": "categoria",
    "modulo ok?": "modulo_ok",
    "area ok?": "area_ok",
    "padrao titulo?": "padrao_titulo",
    "padrao completo?": "padrao_completo",
    "confianca area": "confianca_area",
    "situacao analise": "situacao_analise",
    "situacao": "situacao_analise",
    "desenvolvedor futuro": "desenvolvedor_futuro",
    "observacao geral": "observacao_geral",
    "chamado": "chamado",
    "priorizar": "priorizar",
    "epico": "epico",
    "epico/atividade": "epico",
    "epico atividade": "epico",
    "faixa de idade": "faixa_idade",
    "faixa idade": "faixa_idade",
}

INT_FIELDS = {
    "lead_time_dias",
    "idade_dias",
    "dev_commits",
    "gitlab_mrs",
    "ano_criacao",
    "gitlab_iid",
}
BOOL_FIELDS = {"aberto", "fechado", "sla_mais_90_dias"}
DATETIME_FIELDS = {
    "criado_em",
    "fechado_em",
    "dev_ultimo_commit",
    "mes_criacao",
    "mes_fechamento",
}


def _utc_now() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _normalize_header(name: Optional[str]) -> str:
    if not name:
        return ""
    text = str(name).strip().lower()
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


def _excel_path(explicit: Optional[Path]) -> Path:
    if explicit:
        return explicit
    if _config:
        return Path(_config.EXCEL_OUTPUT)
    return Path(__file__).resolve().parent.parent / "MGI_Dashboard.xlsx"


def _git_data_path() -> Path:
    if _config:
        return Path(_config.GIT_DATA_JSON)
    return Path(__file__).resolve().parent.parent / "gitlab_git_data.json"


def _resolve_layout(ws: Worksheet) -> Tuple[int, int, Dict[str, int]]:
    header_row = 2
    for row in range(1, min(6, ws.max_row + 1)):
        labels = {
            _normalize_header(ws.cell(row=row, column=col).value)
            for col in range(1, ws.max_column + 1)
        }
        labels.discard("")
        if "titulo" in labels or "modulo" in labels:
            header_row = row
            break

    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        norm = _normalize_header(ws.cell(row=header_row, column=col).value)
        if norm and norm not in header_map:
            header_map[norm] = col

    return header_row, header_row + 1, header_map


def _field_column_map(header_map: Dict[str, int]) -> Dict[str, int]:
    """db_col -> indice 0-based na linha do Excel."""
    mapping: Dict[str, int] = {}
    for header_norm, db_col in COLUMN_MAP.items():
        col = header_map.get(header_norm)
        if col is not None:
            mapping[db_col] = col - 1
    return mapping


def _row_value(row: Sequence[Any], index: Optional[int]) -> Any:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def _to_bool(value: Any) -> Optional[bool]:
    if value is None or value == "":
        return None
    text = str(value).strip().lower()
    if text in {"sim", "yes", "true", "1", "s"}:
        return True
    if text in {"nao", "não", "no", "false", "0", "n"}:
        return False
    return None


def _to_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _to_datetime(value: Any) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day).isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).isoformat()
        except ValueError:
            continue
    return text


def _convert_field(db_col: str, raw: Any) -> Any:
    if db_col in BOOL_FIELDS:
        return _to_bool(raw)
    if db_col in INT_FIELDS:
        return _to_int(raw)
    if db_col in DATETIME_FIELDS:
        return _to_datetime(raw)
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        return _to_datetime(raw)
    return str(raw).strip()


def _build_issue_row(
    row: Sequence[Any],
    *,
    iid_index: int,
    repo_index: Optional[int],
    field_cols: Dict[str, int],
    synced_at: str,
) -> Optional[Dict[str, Any]]:
    iid = _to_int(_row_value(row, iid_index))
    if not iid:
        return None

    repo_raw = _row_value(row, repo_index)
    repo = str(repo_raw).strip() if repo_raw not in (None, "") else "contratos_v2"
    if not repo:
        repo = "contratos_v2"

    record: Dict[str, Any] = {
        "issue_key": f"{repo}:{iid}",
        "gitlab_repo": repo,
        "gitlab_iid": iid,
        "synced_at": synced_at,
        "updated_at": synced_at,
    }

    for db_col, col_index in field_cols.items():
        if db_col in {"gitlab_iid"}:
            continue
        record[db_col] = _convert_field(db_col, _row_value(row, col_index))

    return record


def _extract_issue_rows(ws: Worksheet) -> List[Dict[str, Any]]:
    _, data_start, header_map = _resolve_layout(ws)
    field_cols = _field_column_map(header_map)

    iid_col = header_map.get("id") or header_map.get("#")
    if not iid_col:
        raise SystemExit("Coluna ID ausente na aba Dados")

    iid_index = iid_col - 1
    repo_col = header_map.get("repositorio")
    repo_index = repo_col - 1 if repo_col else None

    synced_at = _utc_now()
    rows: List[Dict[str, Any]] = []
    max_col = ws.max_column

    print(f"OK - Lendo linhas {data_start}..{ws.max_row} ({max_col} colunas)")

    for row_num, row in enumerate(
        ws.iter_rows(
            min_row=data_start,
            max_row=ws.max_row,
            min_col=1,
            max_col=max_col,
            values_only=True,
        ),
        start=data_start,
    ):
        record = _build_issue_row(
            row,
            iid_index=iid_index,
            repo_index=repo_index,
            field_cols=field_cols,
            synced_at=synced_at,
        )
        if record:
            rows.append(record)

        if row_num % 500 == 0 and row_num > data_start:
            print(f"OK - {len(rows)} issues lidas (linha {row_num})")

    return rows


def _dedupe_issue_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Remove issue_key repetida no lote (Postgres rejeita upsert com duplicata no mesmo comando)."""
    seen: Dict[str, Dict[str, Any]] = {}
    duplicates: List[str] = []

    for record in rows:
        key = str(record["issue_key"])
        if key in seen:
            duplicates.append(key)
        seen[key] = record

    if duplicates:
        unique_dupes = sorted(set(duplicates))
        sample = ", ".join(unique_dupes[:8])
        suffix = f" (+{len(unique_dupes) - 8} outras)" if len(unique_dupes) > 8 else ""
        print(
            f"AVISO - {len(duplicates)} linhas duplicadas no Excel "
            f"({len(unique_dupes)} issue_keys). Mantida a ultima ocorrencia. "
            f"Ex.: {sample}{suffix}"
        )

    return list(seen.values())


def _dedupe_releases(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen: Dict[tuple[str, str], Dict[str, Any]] = {}
    for record in rows:
        key = (str(record["repositorio"]), str(record["versao"]))
        seen[key] = record
    return list(seen.values())


class SupabaseSync:
    def __init__(self, url: str, service_key: str) -> None:
        self.base = url.rstrip("/") + "/rest/v1"
        self.headers = {
            "apikey": service_key,
            "Authorization": f"Bearer {service_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        }

    def upsert_issues(self, rows: List[Dict[str, Any]]) -> int:
        if not rows:
            return 0
        total = 0
        chunk_size = 200
        for start in range(0, len(rows), chunk_size):
            chunk = rows[start : start + chunk_size]
            response = requests.post(
                f"{self.base}/issues?on_conflict=issue_key",
                headers=self.headers,
                json=chunk,
                timeout=120,
            )
            if not response.ok:
                detail = response.text[:500]
                raise RuntimeError(
                    f"Erro Supabase issues ({response.status_code}): {detail}"
                )
            total += len(chunk)
            print(f"OK - Enviadas {total}/{len(rows)} issues")
        return total

    def upsert_releases(self, rows: List[Dict[str, Any]]) -> int:
        if not rows:
            return 0
        response = requests.post(
            f"{self.base}/releases?on_conflict=repositorio,versao",
            headers=self.headers,
            json=rows,
            timeout=60,
        )
        if not response.ok:
            detail = response.text[:500]
            raise RuntimeError(
                f"Erro Supabase releases ({response.status_code}): {detail}"
            )
        return len(rows)

    def start_sync_run(self) -> str:
        response = requests.post(
            f"{self.base}/sync_runs",
            headers={**self.headers, "Prefer": "return=representation"},
            json={"source": "excel", "status": "running"},
            timeout=30,
        )
        if not response.ok:
            detail = response.text[:500]
            raise RuntimeError(
                f"Erro Supabase sync_runs ({response.status_code}): {detail}"
            )
        return response.json()[0]["id"]

    def finish_sync_run(
        self,
        run_id: str,
        *,
        status: str,
        rows: int,
        releases: int,
        message: str = "",
    ) -> None:
        payload = {
            "status": status,
            "rows_upserted": rows,
            "releases_upserted": releases,
            "finished_at": _utc_now(),
            "message": message[:500],
        }
        response = requests.patch(
            f"{self.base}/sync_runs?id=eq.{run_id}",
            headers=self.headers,
            json=payload,
            timeout=30,
        )
        response.raise_for_status()


def _load_releases(path: Path) -> List[Dict[str, Any]]:
    if not path.exists():
        return []

    with open(path, encoding="utf-8") as handle:
        data = json.load(handle)

    rows: List[Dict[str, Any]] = []
    synced_at = _utc_now()

    def append(repo_name: str, rel: Dict[str, Any]) -> None:
        versao = str(rel.get("versao", "")).strip()
        if not versao:
            return
        rows.append(
            {
                "repositorio": repo_name,
                "versao": versao,
                "data_release": rel.get("data") or None,
                "rotulo": f"{repo_name}: {versao}",
                "synced_at": synced_at,
            }
        )

    if isinstance(data.get("repositorios"), list):
        for repo_block in data["repositorios"]:
            repo_name = repo_block.get("repositorio", "contratos_v2")
            for rel in repo_block.get("releases") or []:
                append(repo_name, rel)
        return rows

    repo_name = data.get("repositorio", "contratos_v2")
    for rel in data.get("releases") or []:
        append(repo_name, rel)
    return rows


def _load_dotenv() -> None:
    """Carrega variaveis de .env na raiz do projeto (se existir)."""
    candidates = [
        Path(__file__).resolve().parent.parent / ".env",
        Path.cwd() / ".env",
        Path.cwd().parent / ".env",
    ]
    for path in candidates:
        if not path.exists():
            continue
        for line in path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            key = key.strip()
            value = value.strip().strip('"').strip("'")
            if key:
                os.environ[key] = value
        print(f"OK - Variaveis carregadas de {path}")
        return


def sync_excel_to_supabase(excel_path: Path, include_releases: bool = True) -> int:
    _load_dotenv()
    url = os.environ.get("SUPABASE_URL", "").strip()
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if not url or not key:
        raise SystemExit(
            "Defina SUPABASE_URL e SUPABASE_SERVICE_ROLE_KEY em .env na raiz do workspace (mgi-workspace/.env)"
        )

    if not excel_path.exists():
        raise SystemExit(f"Excel nao encontrado: {excel_path}")

    print(f"OK - Carregando {excel_path}")
    wb = load_workbook(excel_path, data_only=True)
    if DADOS_SHEET not in wb.sheetnames:
        raise SystemExit(f"Aba '{DADOS_SHEET}' ausente no Excel")

    ws = wb[DADOS_SHEET]
    rows = _dedupe_issue_rows(_extract_issue_rows(ws))
    wb.close()
    print(f"OK - {len(rows)} issues unicas preparadas")

    client = SupabaseSync(url, key)
    run_id: Optional[str] = None
    try:
        run_id = client.start_sync_run()
    except Exception as exc:
        print(f"AVISO - sync_runs indisponivel ({exc}). Continuando upsert...")

    try:
        upserted = client.upsert_issues(rows)
        release_count = 0
        if include_releases:
            release_rows = _dedupe_releases(_load_releases(_git_data_path()))
            release_count = client.upsert_releases(release_rows)
            print(f"OK - {release_count} releases sincronizadas")

        if run_id:
            client.finish_sync_run(
                run_id,
                status="success",
                rows=upserted,
                releases=release_count,
                message=f"sync from {excel_path.name}",
            )
        print(f"OK - Sync concluido: {upserted} issues")
        return upserted
    except Exception as exc:
        if run_id:
            client.finish_sync_run(
                run_id,
                status="error",
                rows=0,
                releases=0,
                message=str(exc),
            )
        raise


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync Excel Dados -> Supabase")
    parser.add_argument("--excel", type=Path, default=None)
    parser.add_argument("--sem-releases", action="store_true")
    args = parser.parse_args()

    sync_excel_to_supabase(_excel_path(args.excel), include_releases=not args.sem_releases)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
