#!/usr/bin/env python3
"""Sincroniza aba Listas com taxonomia oficial (modulos, categorias, areas)."""

from __future__ import annotations

from typing import List, Sequence, Tuple

from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook
from openpyxl.workbook.defined_name import DefinedName

from taxonomy import (
    CANONICAL_MODULES,
    MODULE_CATEGORIES,
    all_accepted_modules_sorted,
    all_module_category_rows,
    all_module_de_para_rows,
    all_standard_areas_sorted,
)

LISTAS_SHEET = "Listas"

MODULO_LIST_COL = 2
MODULO_CAT_LABEL_COL = 17  # Q
MODULO_CAT_VALUE_COL = 18  # R
AREA_PADRAO_COL = 19  # S
CATEGORIA_LIST_COL = 20  # T
MODULO_DEPARA_TAG_COL = 24  # X
MODULO_DEPARA_CANON_COL = 25  # Y


def _write_defined_name(wb: Workbook, name: str, ref: str) -> None:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names.add(DefinedName(name, attr_text=ref))


def sync_listas_taxonomia(wb: Workbook) -> dict:
    """Atualiza listas de modulos, categorias e areas padrao."""
    ws = wb[LISTAS_SHEET]

    # Modulos canonicos (12) — filtros e KPI estrito
    ws.cell(row=1, column=MODULO_LIST_COL).value = "Módulo"
    ws.cell(row=2, column=MODULO_LIST_COL).value = "Todos"
    for idx, module in enumerate(CANONICAL_MODULES, start=3):
        ws.cell(row=idx, column=MODULO_LIST_COL).value = module
    modulo_last = 2 + len(CANONICAL_MODULES)
    mod_col = get_column_letter(MODULO_LIST_COL)
    _write_defined_name(
        wb,
        "Lista_Modulos_Canonico",
        f"{LISTAS_SHEET}!${mod_col}$3:${mod_col}${modulo_last}",
    )

    # De-para tag -> canonico (12 modulos do diagnostico)
    de_para = all_module_de_para_rows()
    ws.cell(row=1, column=MODULO_DEPARA_TAG_COL).value = "De-Para Módulo"
    ws.cell(row=2, column=MODULO_DEPARA_TAG_COL).value = "Tag/Alias"
    ws.cell(row=2, column=MODULO_DEPARA_CANON_COL).value = "Canônico"
    for idx, (source, target) in enumerate(de_para, start=3):
        ws.cell(row=idx, column=MODULO_DEPARA_TAG_COL).value = source
        ws.cell(row=idx, column=MODULO_DEPARA_CANON_COL).value = target
    de_last = 2 + len(de_para)
    x_col = get_column_letter(MODULO_DEPARA_TAG_COL)
    y_col = get_column_letter(MODULO_DEPARA_CANON_COL)
    _write_defined_name(
        wb,
        "Lista_Modulo_DePara",
        f"{LISTAS_SHEET}!${x_col}$3:${y_col}${de_last}",
    )
    _write_defined_name(
        wb,
        "Lista_Modulos_Aceitos",
        f"{LISTAS_SHEET}!${x_col}$3:${x_col}${de_last}",
    )

    module_rows = all_module_category_rows()
    ws.cell(row=1, column=MODULO_CAT_LABEL_COL).value = "Módulo (taxonomia)"
    ws.cell(row=1, column=MODULO_CAT_VALUE_COL).value = "Categoria"
    ws.cell(row=2, column=MODULO_CAT_LABEL_COL).value = "Módulo"
    ws.cell(row=2, column=MODULO_CAT_VALUE_COL).value = "Categoria"
    for idx, (module, category) in enumerate(module_rows, start=3):
        ws.cell(row=idx, column=MODULO_CAT_LABEL_COL).value = module
        ws.cell(row=idx, column=MODULO_CAT_VALUE_COL).value = category
    cat_last = 2 + len(module_rows)
    q_col = get_column_letter(MODULO_CAT_LABEL_COL)
    r_col = get_column_letter(MODULO_CAT_VALUE_COL)
    _write_defined_name(
        wb,
        "Lista_Modulo_Categoria",
        f"{LISTAS_SHEET}!${q_col}$3:${r_col}${cat_last}",
    )

    # Categorias unicas (col T)
    categories = sorted(set(MODULE_CATEGORIES.values()) | {"Operations", "Não mapeado"})
    ws.cell(row=1, column=CATEGORIA_LIST_COL).value = "Categoria Funcional"
    ws.cell(row=2, column=CATEGORIA_LIST_COL).value = "Todos"
    for idx, cat in enumerate(categories, start=3):
        ws.cell(row=idx, column=CATEGORIA_LIST_COL).value = cat
    cat_col = get_column_letter(CATEGORIA_LIST_COL)
    cat_list_last = 2 + len(categories)
    _write_defined_name(
        wb,
        "Lista_Categoria",
        f"{LISTAS_SHEET}!${cat_col}$2:${cat_col}${cat_list_last}",
    )

    # Areas padrao (col S)
    areas = all_standard_areas_sorted()
    ws.cell(row=1, column=AREA_PADRAO_COL).value = "Áreas Padrão"
    ws.cell(row=2, column=AREA_PADRAO_COL).value = "Todos"
    for idx, area in enumerate(areas, start=3):
        ws.cell(row=idx, column=AREA_PADRAO_COL).value = area
    area_col = get_column_letter(AREA_PADRAO_COL)
    area_last = 2 + len(areas)
    _write_defined_name(
        wb,
        "Lista_Areas_Padrao",
        f"{LISTAS_SHEET}!${area_col}$3:${area_col}${area_last}",
    )

    return {
        "modulos_canonicos": len(CANONICAL_MODULES),
        "de_para": len(de_para),
        "categorias": len(categories),
        "areas_padrao": len(areas),
    }
