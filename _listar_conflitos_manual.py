#!/usr/bin/env python3
"""Lista tags cujo mapeamento do pipeline difere da planilha manual."""
from __future__ import annotations

import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from taxonomy import CUSTOM_BUCKET, NON_MODULE_BUCKET, canonical_or_bucket, normalize_module_to_canonical

MANUAL_PATH = Path(r"D:\MGI-Relatórios\Dados-(ajustes manuais).xlsx")
TAG_RE = re.compile(r"^\[([^\]]+)\]")


def pipeline_module_for_tag(tag: str) -> str:
    bucket = canonical_or_bucket(tag)
    if bucket in (CUSTOM_BUCKET, NON_MODULE_BUCKET):
        return bucket
    return normalize_module_to_canonical(tag) or bucket


def manual_module_for_tag(ctr: Counter) -> str:
    return ctr.most_common(1)[0][0]


def equivalent(manual_mod: str, pipe_mod: str) -> bool:
    if manual_mod == pipe_mod:
        return True
    manual_canon = normalize_module_to_canonical(manual_mod) or manual_mod
    return manual_canon == pipe_mod


def main() -> None:
    wb = load_workbook(MANUAL_PATH, data_only=True)
    ws = wb["Dados"]
    manual_by_tag: dict[str, Counter] = defaultdict(Counter)

    for row in range(2, ws.max_row + 1):
        title = str(ws.cell(row, 2).value or "")
        manual_mod = str(ws.cell(row, 3).value or "").strip()
        match = TAG_RE.match(title)
        if not match or not manual_mod:
            continue
        manual_by_tag[match.group(1).strip()][manual_mod] += 1
    wb.close()

    conflicts = []
    for tag, ctr in manual_by_tag.items():
        manual_mod = manual_module_for_tag(ctr)
        pipe_mod = pipeline_module_for_tag(tag)
        if equivalent(manual_mod, pipe_mod):
            continue
        conflicts.append(
            {
                "tag": tag,
                "manual": manual_mod,
                "count": sum(ctr.values()),
                "pipeline": pipe_mod,
                "alternatives": ctr.most_common(5),
            }
        )

    conflicts.sort(key=lambda x: -x["count"])

    print(f"Tags conflitantes: {len(conflicts)}\n")
    for i, item in enumerate(conflicts, 1):
        alts = ", ".join(
            f"{mod}({n})"
            for mod, n in item["alternatives"]
            if mod != item["manual"]
        )
        alt_text = f"  [alt manual: {alts}]" if alts else ""
        print(
            f"{i:2d}. [{item['tag']}]  "
            f"manual={item['manual']!r} ({item['count']})  ->  "
            f"pipeline={item['pipeline']!r}{alt_text}"
        )


if __name__ == "__main__":
    main()
