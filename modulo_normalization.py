#!/usr/bin/env python3
"""Colunas e normalizacao de modulo (Original / Normalizado -> 12 canonicos)."""

from __future__ import annotations

from typing import Dict, Tuple

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from taxonomy import (
    CUSTOM_BUCKET,
    canonical_or_bucket,
    extract_module_tag,
)

MODULE_COLUMN_FIELDS: Tuple[Tuple[str, str], ...] = (
    ("modulo original", "Módulo Original"),
    ("modulo normalizado", "Módulo Normalizado"),
)


def _normalize_header(value: str) -> str:
    import unicodedata

    text = unicodedata.normalize("NFKD", (value or "").strip())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.casefold()


def _header_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=col).value
        if raw:
            mapping[_normalize_header(str(raw))] = col
    return mapping


def ensure_module_columns(ws: Worksheet, header_row: int) -> Dict[str, int]:
    header_map = _header_map(ws, header_row)
    next_col = ws.max_column + 1
    col_map: Dict[str, int] = {}

    for alias, title in MODULE_COLUMN_FIELDS:
        if alias in header_map:
            col_map[alias] = header_map[alias]
            continue
        ws.cell(row=header_row, column=next_col).value = title
        col_map[alias] = next_col
        header_map[alias] = next_col
        next_col += 1

    return col_map


def apply_module_normalization(
    ws: Worksheet,
    header_row: int,
    data_start_row: int,
    last_row: int,
    module_cols: Dict[str, int],
    *,
    sync_modulo_column: bool = True,
    preserve_filled_module: bool = False,
) -> Dict[str, int]:
    """Preenche Módulo Original / Normalizado; opcionalmente sincroniza coluna Módulo."""
    headers = _header_map(ws, header_row)
    mod_col = headers.get("modulo")
    title_col = headers.get("titulo")
    orig_col = module_cols["modulo original"]
    norm_col = module_cols["modulo normalizado"]

    stats = {
        "linhas": 0,
        "canonicos": 0,
        "custom": 0,
        "vazios": 0,
        "modulo_sincronizado": 0,
        "preservados": 0,
    }

    for row in range(data_start_row, last_row + 1):
        issue_id = ws.cell(row=row, column=1).value
        if issue_id in (None, "", "#"):
            continue

        stats["linhas"] += 1
        title = str(ws.cell(row=row, column=title_col).value or "") if title_col else ""
        modulo_cell = str(ws.cell(row=row, column=mod_col).value or "") if mod_col else ""

        if preserve_filled_module and modulo_cell:
            stats["preservados"] += 1
            continue

        tag = extract_module_tag(title) or modulo_cell.strip()
        ws.cell(row=row, column=orig_col).value = tag

        normalized = canonical_or_bucket(tag)
        ws.cell(row=row, column=norm_col).value = normalized

        if not tag:
            stats["vazios"] += 1
        elif normalized == CUSTOM_BUCKET:
            stats["custom"] += 1
        else:
            stats["canonicos"] += 1

        if sync_modulo_column and mod_col and normalized and normalized != CUSTOM_BUCKET:
            current = str(ws.cell(row=row, column=mod_col).value or "").strip()
            if current != normalized:
                ws.cell(row=row, column=mod_col).value = normalized
                stats["modulo_sincronizado"] += 1

    return stats


def apply_module_formulas(
    ws: Worksheet,
    header_row: int,
    data_start_row: int,
    last_row: int,
    module_cols: Dict[str, int],
) -> int:
    """Formulas para recalcular Original/Normalizado a partir do titulo."""
    headers = _header_map(ws, header_row)
    title_col = headers.get("titulo")
    mod_col = headers.get("modulo")
    title_l = get_column_letter(title_col) if title_col else "B"
    mod_l = get_column_letter(mod_col) if mod_col else "C"
    orig_col = module_cols["modulo original"]
    norm_col = module_cols["modulo normalizado"]
    orig_l = get_column_letter(orig_col)

    applied = 0
    for row in range(data_start_row, last_row + 1):
        if ws.cell(row=row, column=1).value in (None, "", "#"):
            continue
        ws.cell(row=row, column=orig_col).value = (
            f'=IF(ISNUMBER(FIND("[",{title_l}{row})),'
            f'MID({title_l}{row},FIND("[",{title_l}{row})+1,'
            f'FIND("]",{title_l}{row})-FIND("[",{title_l}{row})-1),'
            f'{mod_l}{row})'
        )
        ws.cell(row=row, column=norm_col).value = (
            f'=IF({orig_l}{row}="","",'
            f'IFERROR(VLOOKUP({orig_l}{row},Lista_Modulo_DePara,2,FALSE),"{CUSTOM_BUCKET}"))'
        )
        applied += 1
    return applied
